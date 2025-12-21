# portal/datahub_validation.py
from __future__ import annotations

import csv
import io
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from django.core.files.uploadedfile import UploadedFile

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None


# -----------------------------
# Helpers / parsing
# -----------------------------

_HEADER_SEP_RE = re.compile(r"[\s\-]+")
_ALLOWED_DATE_FORMATS = (
    "%Y-%m-%d",      # 2025-12-20
    "%d.%m.%Y",      # 20.12.2025
    "%d/%m/%Y",      # 20/12/2025
    "%m/%d/%Y",      # 12/20/2025
    "%d-%m-%Y",      # 20-12-2025
    "%m-%d-%Y",      # 12-20-2025
)

def normalize_header(h: str) -> str:
    """
    Normalize incoming column names to a stable key:
    - trim
    - lower
    - spaces/dashes -> underscore
    - remove surrounding punctuation
    """
    if h is None:
        return ""
    h = str(h).strip().lower()
    h = h.replace("\ufeff", "")  # BOM if any
    h = _HEADER_SEP_RE.sub("_", h)
    h = re.sub(r"[^\w_]", "", h)
    return h.strip("_")


def parse_decimal(value: Any) -> Optional[Decimal]:
    """
    Accept:
    - Decimal / int / float
    - strings with "." or "," decimal separator
    - empty -> None
    """
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        # float -> str to avoid binary artifacts
        return Decimal(str(value))
    s = str(value).strip()
    if s == "" or s == "—" or s == "-":
        return None

    # normalize thousand separators and decimal commas:
    # examples:
    #  "1,234.56" -> "1234.56"
    #  "1 234,56" -> "1234.56"
    s = s.replace(" ", "")
    # if both separators exist, assume "," thousands when "." decimal (common EN)
    if "," in s and "." in s:
        s = s.replace(",", "")
    else:
        # otherwise treat comma as decimal separator
        s = s.replace(",", ".")

    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError(f"Invalid decimal: {value!r}")


def parse_date(value: Any) -> Optional[date]:
    """
    Accept:
    - date / datetime
    - excel date (openpyxl gives datetime/date already)
    - strings in common formats
    - empty -> None
    """
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    s = str(value).strip()
    if s == "" or s == "—" or s == "-":
        return None

    for fmt in _ALLOWED_DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    # last resort: try ISO-like datetime
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        raise ValueError(f"Invalid date: {value!r}")


def parse_str(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return s


def parse_email(value: Any) -> str:
    s = parse_str(value)
    return s


def parse_url(value: Any) -> str:
    s = parse_str(value)
    return s


# -----------------------------
# Validation engine
# -----------------------------

@dataclass
class ValidationErrorItem:
    row: int                 # 1-based (excluding header)
    column: str
    message: str


@dataclass
class ValidationResult:
    clean_rows: List[Dict[str, Any]]
    errors: List[ValidationErrorItem]
    warnings: List[str]
    headers_in_file: List[str]


Converter = Callable[[Any], Any]

@dataclass
class DatasetSpec:
    key: str
    label: str
    # map normalized input headers -> internal field name
    header_map: Dict[str, str]
    # converters for internal fields
    converters: Dict[str, Converter]
    required: Tuple[str, ...] = ()
    allow_unknown_columns: bool = True


def validate_rows(
    raw_rows: List[Dict[str, Any]],
    headers_in_file: List[str],
    spec: DatasetSpec,
) -> ValidationResult:
    errors: List[ValidationErrorItem] = []
    warnings: List[str] = []
    clean_rows: List[Dict[str, Any]] = []

    # Normalize headers from file
    normalized_file_headers = [normalize_header(h) for h in headers_in_file if h is not None]
    unknown_cols = []

    # Determine mapped internal fields from file headers
    mapped_internal_fields = set()
    for h_norm in normalized_file_headers:
        if h_norm in spec.header_map:
            mapped_internal_fields.add(spec.header_map[h_norm])
        else:
            if h_norm:
                unknown_cols.append(h_norm)

    if unknown_cols and not spec.allow_unknown_columns:
        warnings.append(f"Unknown columns will be ignored: {', '.join(sorted(set(unknown_cols)))}")

    # Check if required fields are present in file at all (by mapping)
    for req in spec.required:
        if req not in mapped_internal_fields:
            warnings.append(f"Missing required column (or alias) in file: {req}")

    # Row-level validation
    for idx, raw in enumerate(raw_rows, start=1):  # 1-based rows
        clean: Dict[str, Any] = {}

        # map each provided column -> internal field
        for raw_key, raw_val in raw.items():
            k_norm = normalize_header(raw_key)
            if not k_norm:
                continue
            internal = spec.header_map.get(k_norm)
            if not internal:
                continue

            conv = spec.converters.get(internal, lambda x: x)
            try:
                clean[internal] = conv(raw_val)
            except Exception as e:
                errors.append(ValidationErrorItem(row=idx, column=internal, message=str(e)))

        # required fields not empty
        for req in spec.required:
            v = clean.get(req)
            if v is None or (isinstance(v, str) and v.strip() == ""):
                errors.append(ValidationErrorItem(row=idx, column=req, message="Required value is missing."))

        clean_rows.append(clean)

    return ValidationResult(
        clean_rows=clean_rows,
        errors=errors,
        warnings=warnings,
        headers_in_file=headers_in_file,
    )


# -----------------------------
# File readers (CSV / XLSX)
# -----------------------------

def read_csv(upload: UploadedFile) -> Tuple[List[Dict[str, Any]], List[str]]:
    raw = upload.read()
    # try utf-8-sig (handles BOM), fallback to utf-8, then cp1251
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            text = None
    if text is None:
        # last resort
        text = raw.decode("utf-8", errors="replace")

    buf = io.StringIO(text)
    reader = csv.DictReader(buf)
    headers = reader.fieldnames or []
    rows = [r for r in reader]
    return rows, headers


def read_xlsx(upload: UploadedFile, sheet: Optional[str] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed. Install it to support Excel imports.")
    wb = openpyxl.load_workbook(upload, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    # first row = headers
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    data_rows: List[Dict[str, Any]] = []

    for r in rows_iter:
        # skip fully empty
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            item[h] = r[i] if i < len(r) else None
        data_rows.append(item)

    return data_rows, headers


def read_tabular_file(upload: UploadedFile) -> Tuple[List[Dict[str, Any]], List[str]]:
    name = upload.name or ""
    ext = os.path.splitext(name)[1].lower()

    if ext in (".csv",):
        return read_csv(upload)
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return read_xlsx(upload)

    raise ValueError("Unsupported file type. Please upload CSV or Excel (.xlsx).")


# -----------------------------
# Dataset specs (start with Cost Centers; extend)
# -----------------------------

COST_CENTERS_SPEC = DatasetSpec(
    key="cost_centers",
    label="Cost centers",
    header_map={
        # accept multiple aliases:
        "code": "code",
        "cost_center_code": "code",
        "costcentre_code": "code",
        "cost_center": "code",
        "name": "name",
        "cost_center_name": "name",
        "business_unit": "business_unit",
        "bu": "business_unit",
        "region": "region",
        "default_approver": "default_approver",         # optional (username/email later)
        "default_approver_username": "default_approver",
    },
    converters={
        "code": parse_str,
        "name": parse_str,
        "business_unit": parse_str,
        "region": parse_str,
        "default_approver": parse_str,
    },
    required=("code", "name"),
    allow_unknown_columns=True,
)
