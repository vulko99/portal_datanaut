from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime, date, timedelta
from collections import defaultdict

from django.contrib.auth.models import User
from .models import Invoice, InvoiceLine, Service, Vendor, Contract, CostCenter
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.db import transaction, IntegrityError
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models.functions import ExtractYear
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, Http404, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.urls import reverse
from urllib.parse import urlencode
from django.views.decorators.http import require_POST
from django.core.exceptions import FieldDoesNotExist

from .models import (
    Vendor,
    Service,
    CostCenter,
    UserProfile,
    Contract,
    Invoice,
    ServiceAssignment,
    ProvisioningRequest,
    InvoiceLine,
    UserProfile,

)
from .forms import ContractUploadForm, InvoiceUploadForm, VendorCreateForm

User = get_user_model()


# -------------------------
# Audit Log (safe integration)
# -------------------------

def _get_audit_model():
    try:
        from .models import AuditEvent
        return AuditEvent
    except Exception:
        return None


def _audit_actor_display(user) -> str:
    try:
        if not user:
            return "—"
        prof = getattr(user, "profile", None)
        full_name = getattr(prof, "full_name", "") if prof else ""
        full_name = (full_name or "").strip()
        if full_name:
            return full_name
        fn = (getattr(user, "get_full_name", lambda: "")() or "").strip()
        if fn:
            return fn
        return getattr(user, "username", "") or getattr(user, "email", "") or "User"
    except Exception:
        return "User"


def _vendor_snapshot(vendor: Vendor) -> dict:
    return {
        "name": vendor.name or "",
        "vendor_type": vendor.vendor_type or "",
        "primary_contact_name": vendor.primary_contact_name or "",
        "primary_contact_email": vendor.primary_contact_email or "",
        "website": vendor.website or "",
        "tags": vendor.tags or "",
        "notes": vendor.notes or "",
        "is_active": bool(getattr(vendor, "is_active", True)),
    }


def _service_snapshot(service: Service) -> dict:
    return {
        "vendor": str(getattr(service.vendor, "name", "") or ""),
        "name": service.name or "",
        "category": service.category or "",
        "service_code": getattr(service, "service_code", "") or "",
        "default_currency": getattr(service, "default_currency", "") or "",
        "default_billing_frequency": getattr(service, "default_billing_frequency", "") or "",
        "owner_display": getattr(service, "owner_display", "") or "",
        "allocation_split": getattr(service, "allocation_split", "") or "",
        "list_price": str(service.list_price) if service.list_price is not None else "",
        "primary_contract": str(service.primary_contract) if service.primary_contract else "",
        "is_active": bool(getattr(service, "is_active", True)),
    }


def _user_snapshot(user_obj: User, profile: UserProfile | None = None) -> dict:
    prof = profile or getattr(user_obj, "profile", None)
    return {
        "username": getattr(user_obj, "username", "") or "",
        "email": getattr(user_obj, "email", "") or "",
        "first_name": getattr(user_obj, "first_name", "") or "",
        "last_name": getattr(user_obj, "last_name", "") or "",
        "full_name": getattr(prof, "full_name", "") if prof else "",
        "cost_center": str(getattr(prof, "cost_center", "") or "") if prof else "",
        "manager": str(getattr(prof, "manager", "") or "") if prof else "",
        "location": getattr(prof, "location", "") if prof else "",
        "legal_entity": getattr(prof, "legal_entity", "") if prof else "",
        "phone_number": getattr(prof, "phone_number", "") if prof else "",
        "is_active": bool(getattr(user_obj, "is_active", True)),
    }


def _diff_snapshots(before: dict, after: dict) -> list[str]:
    field_labels = {
        "name": "Name",
        "vendor_type": "Vendor type",
        "primary_contact_name": "Primary contact name",
        "primary_contact_email": "Primary contact email",
        "website": "Website",
        "tags": "Tags",
        "notes": "Internal notes",
        "vendor": "Vendor",
        "category": "Category",
        "service_code": "Service code",
        "default_currency": "Default currency",
        "default_billing_frequency": "Default billing frequency",
        "owner_display": "Owner",
        "allocation_split": "Allocation split",
        "list_price": "List price",
        "primary_contract": "Primary contract",
        "username": "Username",
        "email": "Email",
        "first_name": "First name",
        "last_name": "Last name",
        "full_name": "Full name",
        "cost_center": "Cost center",
        "manager": "Manager",
        "location": "Location",
        "legal_entity": "Legal entity",
        "phone_number": "Phone number",
        "is_active": "Status",
    }

    def _status_disp(v):
        if v is None:
            return "—"
        return "Active" if bool(v) else "Closed"

    changes: list[str] = []
    keys = set(before.keys()) | set(after.keys())
    for k in sorted(keys):
        old = before.get(k, "")
        new = after.get(k, "")

        if k == "is_active":
            if bool(old) != bool(new):
                label = field_labels.get(k, k)
                changes.append(f"{label}: {_status_disp(old)} → {_status_disp(new)}")
            continue

        old_s = (old or "").strip() if isinstance(old, str) else str(old or "").strip()
        new_s = (new or "").strip() if isinstance(new, str) else str(new or "").strip()
        if old_s != new_s:
            label = field_labels.get(k, k)
            old_disp = old_s if old_s else "—"
            new_disp = new_s if new_s else "—"
            changes.append(f"{label}: {old_disp} → {new_disp}")
    return changes


def _audit_log_event(*, request, object_type: str, object_id: int, description: str, action: str | None = None) -> None:
    AuditEvent = _get_audit_model()
    if not AuditEvent:
        return

    try:
        actor = (request.user if getattr(request, "user", None) and request.user.is_authenticated else None)
        AuditEvent.objects.create(
            object_type=object_type,
            object_id=object_id,
            occurred_at=timezone.now(),
            actor=actor,
            actor_display=_audit_actor_display(actor) if actor else "—",
            description=description,
            action=action,
        )
    except Exception:
        return


def _audit_fetch_events(*, object_type: str, object_id: int, limit: int = 50) -> list:
    AuditEvent = _get_audit_model()
    if not AuditEvent:
        return []

    try:
        return list(
            AuditEvent.objects
            .filter(object_type=object_type, object_id=object_id)
            .order_by("-occurred_at", "-id")[:limit]
        )
    except Exception:
        return []


# -------------------------
# Helpers: parsing / export
# -------------------------

_HEADER_SEP_RE = re.compile(r"[\s\-]+")


def _normalize_header(h: str) -> str:
    h = (h or "").replace("\ufeff", "").strip().lower()
    h = _HEADER_SEP_RE.sub("_", h)
    h = re.sub(r"[^\w_]", "", h)
    return h.strip("_")


def _as_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _parse_date(value) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    s = _as_str(value)
    if not s:
        return None

    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        pass

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Invalid date format: {s}. Use YYYY-MM-DD.")


def _parse_decimal(value) -> Decimal | None:
    s = _as_str(value)
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        raise ValueError(f"Invalid decimal value: {s}")


def _parse_int(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value

    s = _as_str(value)
    if not s:
        return None

    try:
        return int(Decimal(s.replace(",", ".")))
    except Exception:
        raise ValueError(f"Invalid integer value: {s}")


def _detect_format(request, filename: str | None = None) -> str:
    fmt = (request.GET.get("format") or "").lower().strip()
    if fmt in ("csv", "xlsx"):
        return fmt

    if filename:
        fn = filename.lower()
        if fn.endswith(".xlsx"):
            return "xlsx"
        return "csv"

    return "csv"


def _read_csv(uploaded_file) -> list[dict]:
    raw = uploaded_file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1251", errors="replace")

    f = io.StringIO(text)
    reader = csv.DictReader(f)
    rows: list[dict] = []
    for r in reader:
        rows.append({_normalize_header(k): v for k, v in (r or {}).items()})
    return rows


def _read_xlsx(uploaded_file) -> list[dict]:
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl is required for XLSX import. Install it and retry.") from e

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    header: list[str] | None = None
    rows: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(row)

        if header is None:
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            header = [_normalize_header(_as_str(v)) for v in values]
            continue

        if all(v is None or str(v).strip() == "" for v in values):
            continue

        record: dict = {}
        for i, key in enumerate(header):
            if not key:
                continue
            record[key] = "" if i >= len(values) or values[i] is None else values[i]
        rows.append(record)

    return rows


def _read_table(uploaded_file, fmt: str) -> list[dict]:
    if fmt == "xlsx":
        return _read_xlsx(uploaded_file)
    return _read_csv(uploaded_file)


def _workbook_response(filename: str, headers: list[str], rows: list[list[str]]) -> HttpResponse:
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl is required for XLSX export. Install it and retry.") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _csv_response(filename: str, headers: list[str], rows: list[list[str]]) -> HttpResponse:
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)

    resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _first_existing_field(model, candidates):
    """
    Utility: връща първото име на поле от candidates, което реално съществува
    в модела. Ако нито едно не съществува, връща None.
    """
    field_names = {
        f.name for f in model._meta.get_fields()
        if hasattr(f, "attname")
    }
    for name in candidates:
        if name in field_names:
            return name
    return None


# -------------------------
# Importers (per entity)
# -------------------------

def _require_columns(rows: list[dict], required: list[str]) -> None:
    if not rows:
        return
    cols = set(rows[0].keys())
    missing = [c for c in required if c not in cols]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")


@transaction.atomic
def _import_vendors(rows: list[dict], request_user) -> dict:
    _require_columns(rows, ["name"])
    created = 0
    updated = 0

    for r in rows:
        name = _as_str(r.get("name"))
        if not name:
            continue

        defaults = {
            "vendor_type": _as_str(r.get("vendor_type")),
            "tags": _as_str(r.get("tags")),
            "primary_contact_name": _as_str(r.get("primary_contact_name")),
            "primary_contact_email": _as_str(r.get("primary_contact_email")),
            "website": _as_str(r.get("website")),
            "notes": _as_str(r.get("notes")),
        }

        obj = Vendor.objects.filter(name__iexact=name).first()
        if obj:
            for k, v in defaults.items():
                if v != "":
                    setattr(obj, k, v)
            obj.name = name
            obj.save()
            updated += 1
        else:
            Vendor.objects.create(name=name, **defaults)
            created += 1

    return {"created": created, "updated": updated}


@transaction.atomic
def _import_cost_centers(rows: list[dict], request_user) -> dict:
    _require_columns(rows, ["code", "name"])
    created = 0
    updated = 0

    for r in rows:
        code = _as_str(r.get("code"))
        name = _as_str(r.get("name"))
        if not code or not name:
            continue

        defaults = {
            "name": name,
            "business_unit": _as_str(r.get("business_unit")),
            "region": _as_str(r.get("region")),
        }

        obj, was_created = CostCenter.objects.update_or_create(
            code=code,
            defaults=defaults,
        )
        if was_created:
            created += 1
        else:
            updated += 1

    return {"created": created, "updated": updated}


@transaction.atomic
def _import_services(rows: list[dict], request_user) -> dict:
    _require_columns(rows, ["vendor_name", "name"])
    created = 0
    updated = 0

    for r in rows:
        vendor_name = _as_str(r.get("vendor_name"))
        name = _as_str(r.get("name"))
        if not vendor_name or not name:
            continue

        vendor = Vendor.objects.filter(name__iexact=vendor_name).first()
        if not vendor:
            raise ValueError(
                f"Vendor not found for service: {vendor_name} (service={name}). Import vendors first."
            )

        defaults = {
            "category": _as_str(r.get("category")),
            "service_code": _as_str(r.get("service_code")),
            "default_currency": _as_str(r.get("default_currency")),
            "default_billing_frequency": _as_str(r.get("default_billing_frequency")),
            "owner_display": _as_str(r.get("owner_display")),
            "allocation_split": _as_str(r.get("allocation_split")),
        }

        lp = r.get("list_price")
        if _as_str(lp):
            defaults["list_price"] = _parse_decimal(lp)

        obj = Service.objects.filter(vendor=vendor, name__iexact=name).first()
        if obj:
            obj.name = name
            for k, v in defaults.items():
                if v is not None and v != "":
                    setattr(obj, k, v)
            obj.save()
            updated += 1
        else:
            Service.objects.create(vendor=vendor, name=name, **defaults)
            created += 1

    return {"created": created, "updated": updated}


@transaction.atomic
def _import_contracts(rows: list[dict], request_user) -> dict:
    _require_columns(rows, ["vendor_name", "contract_name"])
    created = 0
    updated = 0

    for r in rows:
        vendor_name = _as_str(r.get("vendor_name"))
        contract_name = _as_str(r.get("contract_name"))
        if not vendor_name or not contract_name:
            continue

        vendor = Vendor.objects.filter(name__iexact=vendor_name).first()
        if not vendor:
            raise ValueError(
                f"Vendor not found for contract: {vendor_name} (contract={contract_name}). Import vendors first."
            )

        contract_id = _as_str(r.get("contract_id"))
        qs = Contract.objects.filter(owner=request_user, vendor=vendor, contract_name__iexact=contract_name)
        if contract_id:
            qs = qs.filter(contract_id__iexact=contract_id)

        obj = qs.first()

        defaults = {
            "vendor": vendor,
            "contract_name": contract_name,
            "contract_id": contract_id,
            "contract_type": _as_str(r.get("contract_type")),
            "entity": _as_str(r.get("entity")),
            "currency": _as_str(r.get("currency")),
            "status": _as_str(r.get("status")),
            "uploaded_by": request_user,
        }

        av = r.get("annual_value")
        if _as_str(av):
            defaults["annual_value"] = _parse_decimal(av)

        for field in ("start_date", "end_date", "renewal_date"):
            v = r.get(field)
            if _as_str(v):
                defaults[field] = _parse_date(v)

        npd = r.get("notice_period_days")
        nd = r.get("notice_date")

        if _as_str(npd):
            notice_period_days = _parse_int(npd)
            if notice_period_days not in (30, 60, 90, 120):
                raise ValueError(
                    f"Invalid notice_period_days '{_as_str(npd)}' for contract '{contract_name}'. Allowed: 30, 60, 90, 120."
                )
            defaults["notice_period_days"] = notice_period_days

        if _as_str(nd):
            defaults["notice_date"] = _parse_date(nd)

        end_dt = defaults.get("end_date") or (obj.end_date if obj else None)
        notice_dt = defaults.get("notice_date") if "notice_date" in defaults else (obj.notice_date if obj else None)

        if notice_dt and not end_dt:
            raise ValueError(
                f"Contract '{contract_name}': notice_date is set but end_date is missing."
            )
        if notice_dt and end_dt and notice_dt > end_dt:
            raise ValueError(
                f"Contract '{contract_name}': notice_date ({notice_dt}) must be on/before end_date ({end_dt})."
            )

        if obj:
            for k, v in defaults.items():
                if v is not None and v != "":
                    setattr(obj, k, v)
            obj.save()
            updated += 1
        else:
            Contract.objects.create(owner=request_user, **defaults)
            created += 1

    return {"created": created, "updated": updated}


@transaction.atomic
def _import_invoices(rows: list[dict], request_user) -> dict:
    _require_columns(rows, ["vendor_name", "invoice_number", "invoice_date", "currency", "total_amount"])
    created = 0
    updated = 0

    for r in rows:
        vendor_name = _as_str(r.get("vendor_name"))
        invoice_number = _as_str(r.get("invoice_number"))
        invoice_date = r.get("invoice_date")
        currency = _as_str(r.get("currency"))
        total_amount = r.get("total_amount")

        if not vendor_name or not invoice_number:
            continue

        vendor = Vendor.objects.filter(name__iexact=vendor_name).first()
        if not vendor:
            raise ValueError(
                f"Vendor not found for invoice: {vendor_name} (invoice={invoice_number}). Import vendors first."
            )

        contract = None
        contract_name = _as_str(r.get("contract_name"))
        if contract_name:
            contract = Contract.objects.filter(
                owner=request_user, vendor=vendor, contract_name__iexact=contract_name
            ).first()
            if not contract:
                contract = Contract.objects.filter(owner=request_user, contract_name__iexact=contract_name).first()

        defaults = {
            "invoice_date": _parse_date(invoice_date),
            "currency": currency,
            "total_amount": _parse_decimal(total_amount) or Decimal("0"),
            "notes": _as_str(r.get("notes")),
            "contract": contract,
        }

        ta = r.get("tax_amount")
        if _as_str(ta):
            defaults["tax_amount"] = _parse_decimal(ta)

        for field in ("period_start", "period_end"):
            v = r.get(field)
            if _as_str(v):
                defaults[field] = _parse_date(v)

        obj = Invoice.objects.filter(owner=request_user, vendor=vendor, invoice_number__iexact=invoice_number).first()
        if obj:
            for k, v in defaults.items():
                if v is not None and v != "":
                    setattr(obj, k, v)
            obj.invoice_number = invoice_number
            obj.save()
            updated += 1
        else:
            Invoice.objects.create(
                owner=request_user,
                vendor=vendor,
                invoice_number=invoice_number,
                **defaults,
            )
            created += 1

    return {"created": created, "updated": updated}


@transaction.atomic
def _import_users(rows: list[dict], request_user) -> dict:
    """
    Basic users + profiles import.

    Очаквани колони:
      - username (required)
      - email
      - first_name
      - last_name
      - full_name
      - cost_center_code
      - manager_username
      - location
      - legal_entity
      - is_active  (Active/Closed, 1/0, true/false, yes/no и т.н.)
    """
    _require_columns(rows, ["username"])
    created = 0
    updated = 0

    for r in rows:
        username = _as_str(r.get("username"))
        if not username:
            continue

        email = _as_str(r.get("email"))
        first_name = _as_str(r.get("first_name"))
        last_name = _as_str(r.get("last_name"))
        full_name = _as_str(r.get("full_name"))
        cost_center_code = _as_str(r.get("cost_center_code"))
        manager_username = _as_str(r.get("manager_username"))
        location = _as_str(r.get("location"))
        legal_entity = _as_str(r.get("legal_entity"))
        is_active_raw = (_as_str(r.get("is_active")) or "").lower()

        if is_active_raw in ("0", "false", "no", "closed", "inactive"):
            is_active = False
        elif is_active_raw in ("1", "true", "yes", "open", "active"):
            is_active = True
        else:
            # празно или неразпознато -> приемаме Active
            is_active = True

        user = User.objects.filter(username__iexact=username).first()
        if user:
            updated += 1
        else:
            user = User(username=username)
            try:
                user.set_unusable_password()
            except Exception:
                pass
            created += 1

        if email:
            user.email = email
        if first_name:
            user.first_name = first_name
        if last_name:
            user.last_name = last_name
        user.is_active = is_active
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)

        if full_name:
            profile.full_name = full_name

        cc = None
        if cost_center_code:
            cc = CostCenter.objects.filter(code__iexact=cost_center_code).first()
        profile.cost_center = cc

        manager = None
        if manager_username:
            manager = User.objects.filter(username__iexact=manager_username).first()
        profile.manager = manager

        if location:
            profile.location = location
        if legal_entity:
            profile.legal_entity = legal_entity

        profile.save()

    return {"created": created, "updated": updated}


@transaction.atomic
def _import_permissions(rows: list[dict], request_user) -> dict:
    """
    Import за permissions (User × Service).

    Очаквани колони:
      - username
      - vendor_name
      - service_name
    """
    _require_columns(rows, ["username", "vendor_name", "service_name"])
    created = 0
    updated = 0  # няма real "update", просто създаваме, ако липсва

    for r in rows:
        username = _as_str(r.get("username"))
        vendor_name = _as_str(r.get("vendor_name"))
        service_name = _as_str(r.get("service_name"))

        if not (username and vendor_name and service_name):
            continue

        user = User.objects.filter(username__iexact=username).first()
        if not user:
            raise ValueError(f"User not found for permission row (username='{username}').")

        vendor = Vendor.objects.filter(name__iexact=vendor_name).first()
        if not vendor:
            raise ValueError(
                f"Vendor not found for permission row (vendor='{vendor_name}', username='{username}')."
            )

        service = Service.objects.filter(vendor=vendor, name__iexact=service_name).first()
        if not service:
            raise ValueError(
                f"Service not found for permission row "
                f"(vendor='{vendor_name}', service='{service_name}', username='{username}')."
            )

        _, was_created = ServiceAssignment.objects.get_or_create(
            user=user,
            service=service,
            defaults={"assigned_by": request_user},
        )
        if was_created:
            created += 1

    return {"created": created, "updated": updated}


DATA_ENTITIES = {
    "vendors": {
        "label": "Vendors",
        "template_headers": [
            "name", "vendor_type", "tags", "primary_contact_name",
            "primary_contact_email", "website", "notes",
        ],
        "importer": _import_vendors,
        "exporter": lambda user: [
            [
                v.name,
                v.vendor_type or "",
                v.tags or "",
                v.primary_contact_name or "",
                v.primary_contact_email or "",
                v.website or "",
                v.notes or "",
            ]
            for v in Vendor.objects.all().order_by("name")
        ],
    },
    "cost-centers": {
        "label": "Cost centers",
        "template_headers": ["code", "name", "business_unit", "region"],
        "importer": _import_cost_centers,
        "exporter": lambda user: [
            [c.code, c.name, c.business_unit or "", c.region or ""]
            for c in CostCenter.objects.all().order_by("code")
        ],
    },
    "services": {
        "label": "Services",
        "template_headers": [
            "vendor_name", "name", "category", "service_code",
            "default_currency", "default_billing_frequency",
            "owner_display", "list_price", "allocation_split",
        ],
        "importer": _import_services,
        "exporter": lambda user: [
            [
                s.vendor.name,
                s.name,
                s.category or "",
                s.service_code or "",
                s.default_currency or "",
                s.default_billing_frequency or "",
                s.owner_display or "",
                _as_str(s.list_price) if s.list_price is not None else "",
                s.allocation_split or "",
            ]
            for s in Service.objects.select_related("vendor").order_by("vendor__name", "name")
        ],
    },
    "contracts": {
        "label": "Contracts",
        "template_headers": [
            "vendor_name", "contract_name", "contract_id", "contract_type", "entity",
            "annual_value", "currency", "start_date", "end_date", "renewal_date",
            "notice_period_days", "notice_date",
            "status",
        ],
        "importer": _import_contracts,
        "exporter": lambda user: [
            [
                c.vendor.name,
                c.contract_name,
                c.contract_id or "",
                c.contract_type or "",
                c.entity or "",
                _as_str(c.annual_value) if c.annual_value is not None else "",
                c.currency or "",
                _as_str(c.start_date) if c.start_date else "",
                _as_str(c.end_date) if c.end_date else "",
                _as_str(c.renewal_date) if c.renewal_date else "",
                _as_str(c.notice_period_days) if getattr(c, "notice_period_days", None) else "",
                _as_str(c.notice_date) if getattr(c, "notice_date", None) else "",
                c.status or "",
            ]
            for c in Contract.objects.filter(owner=user)
                .select_related("vendor")
                .order_by("-created_at")
        ],
    },
    "invoices": {
        "label": "Invoices",
        "template_headers": [
            "vendor_name", "contract_name", "invoice_number", "invoice_date", "currency",
            "total_amount", "tax_amount", "period_start", "period_end", "notes",
        ],
        "importer": _import_invoices,
        "exporter": lambda user: [
            [
                i.vendor.name,
                i.contract.contract_name if i.contract else "",
                i.invoice_number,
                _as_str(i.invoice_date),
                i.currency,
                _as_str(i.total_amount),
                _as_str(i.tax_amount) if i.tax_amount is not None else "",
                _as_str(i.period_start) if i.period_start else "",
                _as_str(i.period_end) if i.period_end else "",
                i.notes or "",
            ]
            for i in Invoice.objects.filter(owner=user)
                .select_related("vendor", "contract")
                .order_by("-invoice_date", "-id")
        ],
    },

    # ---------- NEW: Users ----------
    "users": {
        "label": "Users",
        "template_headers": [
            "username",
            "email",
            "first_name",
            "last_name",
            "full_name",
            "cost_center_code",
            "manager_username",
            "location",
            "legal_entity",
            "is_active",
        ],
        "importer": _import_users,
        "exporter": lambda user: [
            [
                u.username,
                u.email or "",
                u.first_name or "",
                u.last_name or "",
                getattr(getattr(u, "profile", None), "full_name", "") or "",
                getattr(getattr(getattr(u, "profile", None), "cost_center", None), "code", "") or "",
                getattr(getattr(getattr(u, "profile", None), "manager", None), "username", "") or "",
                getattr(getattr(u, "profile", None), "location", "") or "",
                getattr(getattr(u, "profile", None), "legal_entity", "") or "",
                "Active" if u.is_active else "Closed",
            ]
            for u in User.objects
                .select_related("profile", "profile__cost_center", "profile__manager")
                .order_by("username")
        ],
    },

    # ---------- NEW: Permissions (User · Service) ----------
    "permissions": {
        "label": "Permissions (user · service)",
        "template_headers": [
            "username",
            "vendor_name",
            "service_name",
        ],
        "importer": _import_permissions,
        "exporter": lambda user: [
            [
                a.user.username if a.user else "",
                a.service.vendor.name if a.service and a.service.vendor else "",
                a.service.name if a.service else "",
            ]
            for a in ServiceAssignment.objects
                .select_related("user", "service", "service__vendor")
                .order_by("user__username", "service__vendor__name", "service__name")
        ],
    },
}


def _get_entity_or_404(entity: str) -> dict:
    cfg = DATA_ENTITIES.get(entity)
    if not cfg:
        raise Http404("Unknown Data Hub entity.")
    return cfg



# ---------- 
# DASHBOARD
# ----------

from datetime import date, datetime, timedelta
from decimal import Decimal
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
import json

from .models import Contract, Invoice, Vendor

# ако helper-ът не е в този файл, импортни го от utils
# from .utils import _first_existing_field


@login_required
def dashboard(request):
    user = request.user
    today = date.today()
    last_12m_start = today - timedelta(days=365)
    prev_12m_start = last_12m_start - timedelta(days=365)
    prev_12m_end = last_12m_start - timedelta(days=1)
    renewals_window_end = today + timedelta(days=90)

    # ==========================================================
    # 1) ИНВОЙСИ – общ разход, quarterly chart, top vendors
    # ==========================================================
    amount_field = _first_existing_field(
        Invoice,
        ["total_amount", "amount", "net_amount", "gross_amount", "value"],
    )
    date_field = _first_existing_field(
        Invoice,
        ["invoice_date", "date", "issue_date", "period_start", "period_end"],
    )

    hero_total_spend = Decimal("0")
    hero_spend_change_pct = None
    chart_quarter_labels: list[str] = []
    chart_quarter_actual: list[float] = []
    chart_spend_vendor_labels: list[str] = []
    chart_spend_vendor_values: list[float] = []
    vendor_spend_rows: list[dict] = []

    if amount_field and date_field:
        base_qs = Invoice.objects.filter(owner=user)

        # текущи 12 месеца
        last12_filter = {
            f"{date_field}__gte": last_12m_start,
            f"{date_field}__lte": today,
        }
        last12_qs = base_qs.filter(**last12_filter)

        # предишни 12 месеца (за % промяна)
        prev12_filter = {
            f"{date_field}__gte": prev_12m_start,
            f"{date_field}__lte": prev_12m_end,
        }
        prev12_qs = base_qs.filter(**prev12_filter)

        hero_total_spend = (
            last12_qs.aggregate(total=Sum(amount_field))["total"]
            or Decimal("0")
        )
        prev_total = prev12_qs.aggregate(total=Sum(amount_field))["total"]

        if prev_total not in (None, 0, Decimal("0")):
            hero_spend_change_pct = (
                (hero_total_spend - prev_total) / prev_total * Decimal("100")
            )

        # агрегиране в Python
        quarter_buckets: dict[str, Decimal] = defaultdict(Decimal)
        vendor_buckets: dict[int, Decimal] = defaultdict(Decimal)
        vendor_ids: set[int] = set()

        for inv in last12_qs.select_related("vendor"):
            dt = getattr(inv, date_field, None)
            if isinstance(dt, datetime):
                dt = dt.date()
            if not isinstance(dt, date):
                continue

            # quarter label: 2025-Q1
            q = ((dt.month - 1) // 3) + 1
            q_label = f"{dt.year}-Q{q}"

            amount = getattr(inv, amount_field, None) or Decimal("0")
            if not isinstance(amount, Decimal):
                amount = Decimal(str(amount))

            quarter_buckets[q_label] += amount

            v = getattr(inv, "vendor", None)
            if v:
                vendor_buckets[v.pk] += amount
                vendor_ids.add(v.pk)

        # сортиране на quarter-ите по време
        chart_quarter_labels = sorted(quarter_buckets.keys())
        chart_quarter_actual = [
            float(quarter_buckets[label]) for label in chart_quarter_labels
        ]

        # Top 5 vendors по spend
        top_vendor_pairs = sorted(
            vendor_buckets.items(),
            key=lambda kv: kv[1],
            reverse=True,
        )[:5]

        vendors_by_id = {
            v.pk: v for v in Vendor.objects.filter(id__in=vendor_ids)
        }

        for vid, total in top_vendor_pairs:
            v = vendors_by_id.get(vid)
            if not v:
                continue

            chart_spend_vendor_labels.append(v.name)
            chart_spend_vendor_values.append(float(total))

            vendor_spend_rows.append(
                {
                    "vendor_name": v.name,
                    "category": getattr(v, "vendor_type", "") or "",
                    "total_spend": total,
                }
            )

    # ==========================================================
    # 2) КОНТРАКТИ – hero KPIs, status pie, upcoming renewals
    # ==========================================================
    contracts_qs = (
        Contract.objects.filter(owner=user)
        .select_related("vendor")
        .order_by("renewal_date", "end_date")
    )

    hero_contracts_total = contracts_qs.count()
    hero_contracts_vendors = (
        contracts_qs.values("vendor_id").distinct().count()
    )
    hero_contracts_entities = (
        contracts_qs.values("entity").distinct().count()
    )

    # contracts by status
    status_counts: dict[str, int] = {}
    for c in contracts_qs:
        status = getattr(c, "status", "") or "Unknown"
        status_counts[status] = status_counts.get(status, 0) + 1

    chart_status_labels = list(status_counts.keys())
    chart_status_values = list(status_counts.values())

    # upcoming renewals (90 дни)
    upcoming_rows: list[dict] = []

    for c in contracts_qs:
        d = getattr(c, "renewal_date", None) or getattr(c, "end_date", None)
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date):
            continue
        if not (today <= d <= renewals_window_end):
            continue

        days_to = (d - today).days
        if days_to <= 30:
            risk = "High"
        elif days_to <= 60:
            risk = "Medium"
        else:
            risk = "Low"

        upcoming_rows.append(
            {
                "vendor_name": c.vendor.name if c.vendor else "",
                "contract_name": (
                    getattr(c, "contract_name", "")
                    or getattr(c, "contract_id", "")
                    or ""
                ),
                "entity": getattr(c, "entity", "") or "",
                "renewal_date": d,
                "annual_value": getattr(c, "annual_value", None),
                "currency": getattr(c, "currency", "") or "",
                "risk": risk,
            }
        )

    upcoming_rows.sort(key=lambda r: r["renewal_date"])
    hero_renewals_90d = len(upcoming_rows)

    # ==========================================================
    # CONTEXT – списъците за charts вече са JSON string
    # ==========================================================
    context = {
        # hero
        "hero_total_spend": hero_total_spend,
        "hero_spend_change_pct": hero_spend_change_pct,
        "hero_contracts_total": hero_contracts_total,
        "hero_contracts_vendors": hero_contracts_vendors,
        "hero_contracts_entities": hero_contracts_entities,
        "hero_renewals_90d": hero_renewals_90d,
        # charts (JSON strings за template-а)
        "chart_quarter_labels": json.dumps(chart_quarter_labels),
        "chart_quarter_actual": json.dumps(chart_quarter_actual),
        "chart_spend_vendor_labels": json.dumps(chart_spend_vendor_labels),
        "chart_spend_vendor_values": json.dumps(chart_spend_vendor_values),
        "chart_status_labels": json.dumps(chart_status_labels),
        "chart_status_values": json.dumps(chart_status_values),
        # tables
        "upcoming_rows": upcoming_rows,
        "vendor_spend_rows": vendor_spend_rows,
        # доп. инфо
        "last_12m_start": last_12m_start,
        "today": today,
    }

    return render(request, "portal/dashboard.html", context)






# ----------
# CONTRACTS
# ----------

def _contract_snapshot(contract: Contract) -> dict:
    """
    Snapshot на ключовите полета (за diff в audit лог-а).
    """
    return {
        "vendor_id": contract.vendor_id,
        "contract_name": contract.contract_name or "",
        "contract_id": contract.contract_id or "",
        "contract_type": contract.contract_type or "",
        "entity": contract.entity or "",
        "currency": contract.currency or "",
        "annual_value": str(contract.annual_value) if contract.annual_value is not None else "",
        "start_date": contract.start_date.isoformat() if contract.start_date else "",
        "end_date": contract.end_date.isoformat() if contract.end_date else "",
        "renewal_date": contract.renewal_date.isoformat() if contract.renewal_date else "",
        "notice_period_days": contract.notice_period_days or "",
        "notice_date": contract.notice_date.isoformat() if contract.notice_date else "",
    }


@login_required
def contract_list(request):
    """
    Contract inventory list with:
    - show_closed toggle
    - rows-per-page
    - pagination
    - optional inline selected_contract + audit_events
    - inline edit (Save/Discard) за избрания contract
    """

    # Базов queryset – всички контракти на този user
    base_qs = (
        Contract.objects.filter(owner=request.user)
        .select_related("vendor", "owning_cost_center")
        .order_by("-start_date", "-created_at")
    )

    # Общо контракти (за заглавието)
    total_contracts = base_qs.count()

    # --- Show closed ---
    show_closed_param = request.GET.get("show_closed", "0")
    show_closed = show_closed_param == "1"

    filtered_qs = base_qs
    if not show_closed:
        # Ако имаш други статуси за "затворен", добави ги тук
        closed_statuses = ["expired", "terminated", "cancelled", "closed"]
        filtered_qs = filtered_qs.exclude(status__in=closed_statuses)

    # --- Rows per page ---
    rows_options = [10, 20, 30, 50, 100, 250]
    try:
        rows_per_page = int(request.GET.get("rows") or 50)
    except (TypeError, ValueError):
        rows_per_page = 50
    if rows_per_page <= 0:
        rows_per_page = 50

    # --- Текуща страница ---
    try:
        current_page = int(request.GET.get("page") or 1)
    except (TypeError, ValueError):
        current_page = 1

    paginator = Paginator(filtered_qs, rows_per_page)
    page_obj = paginator.get_page(current_page)
    contracts = list(page_obj.object_list)

    # --- Кой contract е избран за долния панел ---
    selected_contract = None
    selected_id = request.GET.get("selected")
    if selected_id:
        try:
            selected_contract = (
                filtered_qs.select_related("vendor", "owning_cost_center")
                .get(pk=selected_id)
            )
        except Contract.DoesNotExist:
            selected_contract = None

    # ---- POST ---- (add или inline update)
    if request.method == "POST":
        action = _as_str(request.POST.get("action")) or "create"

        # 1) Add contract (modal)
        if action == "create":
            form = ContractUploadForm(request.POST, request.FILES)
            if form.is_valid():
                contract = form.save(owner=request.user, uploaded_by=request.user)

                # AUDIT: create
                _audit_log_event(
                    request=request,
                    object_type="Contract",
                    object_id=contract.pk,
                    action="create",
                    description=f"Created contract '{contract.contract_name}'.",
                )

                messages.success(
                    request,
                    f"Contract '{contract.contract_name}' saved successfully.",
                )
                return redirect("portal:contracts")

        # 2) Inline update на избрания contract в долния панел
        elif action == "update_selected":
            errors: list[str] = []

            selected_id = _as_str(request.POST.get("selected"))
            contract = base_qs.filter(pk=selected_id).first()
            if not contract:
                messages.error(request, "Selected contract was not found.")
                return redirect("portal:contracts")

            before = _contract_snapshot(contract)

            vendor_id = _as_str(request.POST.get("vendor_id"))
            contract_name = _as_str(request.POST.get("contract_name"))
            contract_id = _as_str(request.POST.get("contract_id"))
            contract_type = _as_str(request.POST.get("contract_type"))
            entity = _as_str(request.POST.get("entity"))
            currency = _as_str(request.POST.get("currency"))
            annual_value_raw = _as_str(request.POST.get("annual_value"))
            start_date_raw = _as_str(request.POST.get("start_date"))
            end_date_raw = _as_str(request.POST.get("end_date"))
            renewal_date_raw = _as_str(request.POST.get("renewal_date"))
            notice_period_raw = _as_str(request.POST.get("notice_period_days"))
            notice_date_raw = _as_str(request.POST.get("notice_date"))

            # Задължителни полета
            if not contract_name:
                errors.append("Contract name is required.")

            vendor = None
            if not vendor_id:
                errors.append("Vendor is required.")
            else:
                vendor = Vendor.objects.filter(pk=vendor_id).first()
                if not vendor:
                    errors.append("Selected vendor does not exist.")

            # Decimal
            annual_value = None
            if annual_value_raw:
                try:
                    annual_value = _parse_decimal(annual_value_raw)
                except Exception as e:
                    errors.append(str(e))

            # Дати
            start_date = None
            end_date = None
            renewal_date = None
            notice_date = None
            try:
                if start_date_raw:
                    start_date = _parse_date(start_date_raw)
                if end_date_raw:
                    end_date = _parse_date(end_date_raw)
                if renewal_date_raw:
                    renewal_date = _parse_date(renewal_date_raw)
                if notice_date_raw:
                    notice_date = _parse_date(notice_date_raw)
            except Exception as e:
                errors.append(str(e))

            # Notice period
            notice_period_days = None
            if notice_period_raw:
                try:
                    notice_period_days = _parse_int(notice_period_raw)
                except Exception as e:
                    errors.append(str(e))

            if notice_date and not end_date:
                errors.append("If a notice date is set, end date is required.")
            if notice_date and end_date and notice_date > end_date:
                errors.append("Notice date must be on or before contract end date.")

            if errors:
                for e in errors:
                    messages.error(request, e)
            else:
                contract.vendor = vendor
                contract.contract_name = contract_name
                contract.contract_id = contract_id
                contract.contract_type = contract_type
                contract.entity = entity
                contract.currency = currency
                contract.annual_value = annual_value
                contract.start_date = start_date
                contract.end_date = end_date
                contract.renewal_date = renewal_date
                contract.notice_period_days = notice_period_days
                contract.notice_date = notice_date
                contract.save()

                # AUDIT: update (diff)
                after = _contract_snapshot(contract)
                changes = _diff_snapshots(before, after)
                _audit_log_event(
                    request=request,
                    object_type="Contract",
                    object_id=contract.pk,
                    action="update",
                    description="; ".join(changes) if changes else "Contract updated.",
                )

                messages.success(request, "Contract updated successfully.")

                # >>> важно: запазваме избрания contract + page/rows/show_closed
                page = _as_str(request.POST.get("page")) or "1"
                rows = _as_str(request.POST.get("rows")) or str(rows_per_page)
                show_closed_post = _as_str(request.POST.get("show_closed"))
                if show_closed_post not in ("0", "1"):
                    show_closed_post = "1" if show_closed else "0"

                params = {
                    "page": page,
                    "rows": rows,
                    "show_closed": show_closed_post,
                    "selected": str(contract.pk),
                }
                url = reverse("portal:contracts") + "?" + urlencode(params) + "#contract-details"
                return redirect(url)

        # ако е друго action – просто падаме надолу и ще рендернем страницата
        else:
            form = ContractUploadForm(request.POST, request.FILES)
    else:
        form = ContractUploadForm()

    # Audit events – вече реални, за избрания contract
    audit_events: list = []
    if selected_contract:
        audit_events = _audit_fetch_events(
            object_type="Contract",
            object_id=selected_contract.pk,
            limit=50,
        )

    # Vendors dropdown за inline формата
    vendors = Vendor.objects.all().order_by("name")

    context = {
        "contracts": contracts,
        "form": form,
        "total_contracts": total_contracts,
        "page_obj": page_obj,
        "current_page": page_obj.number,
        "rows_per_page": rows_per_page,
        "rows_options": rows_options,
        "show_closed": show_closed,
        "selected_contract": selected_contract,
        "audit_events": audit_events,
        "vendors": vendors,
    }
    return render(request, "portal/contracts.html", context)


@login_required
def contract_detail(request, pk):
    contract = get_object_or_404(
        Contract.objects.select_related("vendor", "owning_cost_center", "owner"),
        pk=pk,
        owner=request.user,
    )

    invoices = (
        contract.invoices.all()
        .select_related("vendor")
        .order_by("-invoice_date", "-id")
    )

    vendors = Vendor.objects.all().order_by("name")
    cost_centers = CostCenter.objects.all().order_by("code")

    if request.method == "POST":
        action = _as_str(request.POST.get("action")) or "update"

        if action == "delete":
            name = contract.contract_name

            # AUDIT: delete
            _audit_log_event(
                request=request,
                object_type="Contract",
                object_id=contract.pk,
                action="delete",
                description=f"Deleted contract '{name}'.",
            )

            contract.delete()
            messages.success(request, f"Contract '{name}' was deleted.")
            return redirect("portal:contracts")

        errors: list[str] = []
        before = _contract_snapshot(contract)

        vendor_id = _as_str(request.POST.get("vendor_id"))
        contract_name = _as_str(request.POST.get("contract_name"))
        contract_id = _as_str(request.POST.get("contract_id"))
        contract_type = _as_str(request.POST.get("contract_type"))
        entity = _as_str(request.POST.get("entity"))
        cost_center_id = _as_str(request.POST.get("cost_center_id"))
        currency = _as_str(request.POST.get("currency"))
        status = _as_str(request.POST.get("status"))
        annual_value_raw = _as_str(request.POST.get("annual_value"))
        start_date_raw = _as_str(request.POST.get("start_date"))
        end_date_raw = _as_str(request.POST.get("end_date"))
        renewal_date_raw = _as_str(request.POST.get("renewal_date"))
        notice_period_raw = _as_str(request.POST.get("notice_period_days"))
        notice_date_raw = _as_str(request.POST.get("notice_date"))
        notes = _as_str(request.POST.get("notes"))

        if not contract_name:
            errors.append("Contract name is required.")

        vendor = None
        if not vendor_id:
            errors.append("Vendor is required.")
        else:
            vendor = Vendor.objects.filter(pk=vendor_id).first()
            if not vendor:
                errors.append("Selected vendor does not exist.")

        owning_cost_center = None
        if cost_center_id:
            owning_cost_center = CostCenter.objects.filter(pk=cost_center_id).first()
            if not owning_cost_center:
                errors.append("Selected cost centre does not exist.")

        annual_value = None
        if annual_value_raw:
            try:
                annual_value = _parse_decimal(annual_value_raw)
            except Exception as e:
                errors.append(str(e))

        start_date = None
        end_date = None
        renewal_date = None
        notice_date = None
        try:
            if start_date_raw:
                start_date = _parse_date(start_date_raw)
            if end_date_raw:
                end_date = _parse_date(end_date_raw)
            if renewal_date_raw:
                renewal_date = _parse_date(renewal_date_raw)
            if notice_date_raw:
                notice_date = _parse_date(notice_date_raw)
        except Exception as e:
            errors.append(str(e))

        notice_period_days = None
        if notice_period_raw:
            try:
                notice_period_days = _parse_int(notice_period_raw)
            except Exception as e:
                errors.append(str(e))

        if notice_date and not end_date:
            errors.append("If a notice date is set, end date is required.")
        if notice_date and end_date and notice_date > end_date:
            errors.append("Notice date must be on or before contract end date.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            contract.vendor = vendor
            contract.contract_name = contract_name
            contract.contract_id = contract_id
            contract.contract_type = contract_type
            contract.entity = entity
            contract.owning_cost_center = owning_cost_center
            contract.currency = currency
            contract.status = status or contract.status
            contract.annual_value = annual_value
            contract.start_date = start_date
            contract.end_date = end_date
            contract.renewal_date = renewal_date
            contract.notice_period_days = notice_period_days
            contract.notice_date = notice_date
            contract.notes = notes
            contract.save()

            after = _contract_snapshot(contract)
            changes = _diff_snapshots(before, after)
            _audit_log_event(
                request=request,
                object_type="Contract",
                object_id=contract.pk,
                action="update",
                description="; ".join(changes) if changes else "Contract updated.",
            )

            messages.success(request, "Contract updated successfully.")
            return redirect("portal:contract_detail", pk=contract.pk)

    audit_events = _audit_fetch_events(
        object_type="Contract",
        object_id=contract.pk,
        limit=50,
    )

    context = {
        "contract": contract,
        "invoices": invoices,
        "vendors": vendors,
        "cost_centers": cost_centers,
        "audit_events": audit_events,
    }
    return render(request, "portal/contract_detail.html", context)






# ----------
# INVOICING
# ----------

from decimal import Decimal, ROUND_HALF_UP

from django.core.paginator import Paginator
from django.urls import reverse
from django.db.models import Sum
from urllib.parse import urlencode


def _invoice_snapshot(inv):
    """
    Ползва се за audit diff – държим само 'човешките' полета.
    """
    return {
        "vendor": getattr(inv.vendor, "name", None),
        "contract": getattr(inv.contract, "contract_name", None),
        "invoice_number": inv.invoice_number or "",
        "invoice_date": inv.invoice_date.isoformat() if inv.invoice_date else None,
        "currency": inv.currency or "",
        "total_amount": str(inv.total_amount) if inv.total_amount is not None else None,
        "tax_amount": str(inv.tax_amount) if inv.tax_amount is not None else None,
        "period_start": inv.period_start.isoformat() if inv.period_start else None,
        "period_end": inv.period_end.isoformat() if inv.period_end else None,
        "notes": inv.notes or "",
    }


def _render_description(pattern: str, ctx: dict) -> str:
    """
    Проста замяна на плейсхолдъри:
      {service_name}, {username}, {full_name}, {cost_center}
    Оставя непознати плейсхолдъри, без да хвърля грешка.
    """
    if not pattern:
        pattern = "{service_name} – {username}"
    result = pattern
    for key, value in ctx.items():
        token = "{" + key + "}"
        result = result.replace(token, str(value) if value is not None else "")
    return result


@login_required
def invoice_list(request):
    """
    Invoice inventory screen:
    - header stats
    - show_closed toggle (зареждаме го, но засега не филтрираме – няма статус поле)
    - rows-per-page
    - pagination
    - inline selected_invoice + breakdown (cost center / service) + audit
    - Add invoice modal (InvoiceUploadForm)
    - Lines & splits: добавяне / триене на InvoiceLine + generate_from_assignments
    """

    show_closed = (request.GET.get("show_closed") in ("1", "true", "True", "on", "yes"))

    # --- rows per page ---
    rows_options = [10, 20, 30, 50, 100, 250]
    try:
        rows_per_page = int(request.GET.get("rows") or 50)
    except (TypeError, ValueError):
        rows_per_page = 50
    if rows_per_page not in rows_options:
        rows_per_page = 50

    # --- page ---
    raw_page = (request.GET.get("page") or "").strip()
    try:
        page_number = int(raw_page) if raw_page else 1
    except ValueError:
        page_number = 1
    if page_number < 1:
        page_number = 1

    # --- selected invoice id ---
    selected_id = _as_str(request.GET.get("selected"))

    add_form_has_errors = False

    # -------------------------
    # POST: Add (modal) или inline update/delete/add_line/delete_line/generate_from_assignments
    # -------------------------
    if request.method == "POST":
        inline_selected = _as_str(request.POST.get("selected"))

        # 1) INLINE операции върху вече избран invoice
        if inline_selected:
            invoice = (
                Invoice.objects.filter(owner=request.user)
                .select_related("vendor", "contract")
                .filter(pk=inline_selected)
                .first()
            )
            if not invoice:
                messages.error(request, "Selected invoice was not found.")
                return redirect("portal:invoices")

            action = _as_str(request.POST.get("action")) or "update"

            # redirect helper – пазим page/rows/show_closed и selected
            def _redirect_back(include_selected: bool = True):
                post_page = _as_str(request.POST.get("page")) or "1"
                post_rows = _as_str(request.POST.get("rows")) or str(rows_per_page)
                post_show_closed = _as_str(
                    request.POST.get("show_closed") or ("1" if show_closed else "0")
                )
                params = {
                    "page": post_page,
                    "rows": post_rows,
                    "show_closed": post_show_closed,
                }
                if include_selected:
                    params["selected"] = str(invoice.pk)
                url = (
                    reverse("portal:invoices")
                    + "?"
                    + urlencode(params)
                    + "#invoice-details"
                )
                return redirect(url)

            # 1a) DELETE invoice
            if action == "delete":
                number = invoice.invoice_number or f"Invoice {invoice.pk}"

                _audit_log_event(
                    request=request,
                    object_type="Invoice",
                    object_id=invoice.pk,
                    action="delete",
                    description=f"Deleted invoice '{number}'.",
                )

                invoice.delete()
                messages.success(request, f"Invoice '{number}' was deleted.")
                return _redirect_back(include_selected=False)

            # 1b) DELETE line
            if action == "delete_line":
                line_id = _as_str(request.POST.get("line_id"))
                if not line_id:
                    messages.error(request, "Line ID is required.")
                    return _redirect_back(include_selected=True)

                line = InvoiceLine.objects.filter(
                    pk=line_id,
                    invoice=invoice,
                ).first()

                if not line:
                    messages.error(request, "Invoice line was not found.")
                    return _redirect_back(include_selected=True)

                desc = (
                    f"Deleted line #{line.pk}: '{line.description}' "
                    f"amount {line.line_amount} {line.currency or invoice.currency}."
                )

                line.delete()

                _audit_log_event(
                    request=request,
                    object_type="Invoice",
                    object_id=invoice.pk,
                    action="update",  # invoice е променен
                    description=desc,
                )

                messages.success(request, "Invoice line deleted.")
                return _redirect_back(include_selected=True)

            # 1c) GENERATE FROM ASSIGNMENTS
            if action == "generate_from_assignments":
                errors: list[str] = []

                service_id = _as_str(request.POST.get("gen_service_id"))
                description_pattern = _as_str(
                    request.POST.get("gen_description_pattern")
                ) or "{service_name} – {username}"
                use_net = (
                    _as_str(request.POST.get("gen_use_net"))
                    in ("1", "true", "True", "on", "yes")
                )
                clear_existing = (
                    _as_str(request.POST.get("gen_clear_existing"))
                    in ("1", "true", "True", "on", "yes")
                )

                service = None
                if not service_id:
                    errors.append("Service is required to generate splits.")
                else:
                    service = (
                        Service.objects.select_related("vendor")
                        .filter(pk=service_id)
                        .first()
                    )
                    if not service:
                        errors.append("Selected service does not exist.")

                if errors:
                    for e in errors:
                        messages.error(request, e)
                    return _redirect_back(include_selected=True)

                # assignments за този service
                assignments_qs = (
                    ServiceAssignment.objects.filter(service=service)
                    .select_related(
                        "user",
                        "user__profile",
                        "user__profile__cost_center",
                    )
                    .order_by("user__username")
                )
                assignments = list(assignments_qs)

                if not assignments:
                    messages.error(
                        request,
                        "No service assignments found for this service.",
                    )
                    return _redirect_back(include_selected=True)

                # Базова сума – total или net (total - tax)
                total_amount = invoice.total_amount or Decimal("0")
                if use_net and invoice.tax_amount is not None:
                    base_amount = (invoice.total_amount or Decimal("0")) - (
                        invoice.tax_amount or Decimal("0")
                    )
                else:
                    base_amount = total_amount

                if base_amount <= 0:
                    messages.error(
                        request,
                        "Invoice total/net must be positive to generate splits.",
                    )
                    return _redirect_back(include_selected=True)

                # Разпределяме базовата сума равномерно с 2 десетични
                n = len(assignments)
                per_raw = base_amount / Decimal(n)
                per_rounded = per_raw.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

                amounts: list[Decimal] = []
                remaining = base_amount
                for idx in range(n):
                    if idx == n - 1:
                        amt = remaining
                    else:
                        amt = per_rounded
                        remaining -= amt
                    amounts.append(amt)

                # По желание чистим старите линии
                if clear_existing:
                    invoice.lines.all().delete()

                created_count = 0
                for assignment, line_amount in zip(assignments, amounts):
                    user = assignment.user
                    profile = getattr(user, "profile", None)
                    cost_center = getattr(profile, "cost_center", None)

                    full_name = (
                        getattr(profile, "full_name", "") or user.get_full_name() or user.username
                    )
                    cc_label = getattr(cost_center, "code", "")

                    ctx = {
                        "service_name": service.name,
                        "username": user.username,
                        "full_name": full_name,
                        "cost_center": cc_label,
                    }
                    description = _render_description(description_pattern, ctx)

                    InvoiceLine.objects.create(
                        invoice=invoice,
                        service=service,
                        description=description,
                        quantity=Decimal("1"),
                        unit_price=None,
                        line_amount=line_amount,
                        currency=invoice.currency,
                        cost_center=cost_center,
                        user=user,
                    )
                    created_count += 1

                _audit_log_event(
                    request=request,
                    object_type="Invoice",
                    object_id=invoice.pk,
                    action="update",
                    description=(
                        f"Generated {created_count} invoice lines from assignments "
                        f"for service '{service.name}' "
                        f"using {'net' if use_net else 'total'} amount {base_amount}."
                    ),
                )

                messages.success(
                    request,
                    f"Generated {created_count} invoice lines from assignments.",
                )
                return _redirect_back(include_selected=True)

            # 1d) ADD LINE (нов InvoiceLine)
            if action == "add_line":
                errors: list[str] = []

                service_id = _as_str(request.POST.get("line_service_id"))
                description = _as_str(request.POST.get("line_description"))
                quantity_raw = _as_str(request.POST.get("line_quantity")) or "1"
                amount_raw = _as_str(request.POST.get("line_amount"))
                currency = (
                    _as_str(request.POST.get("line_currency"))
                    or invoice.currency
                    or ""
                )
                cost_center_id = _as_str(request.POST.get("line_cost_center_id"))
                user_id = _as_str(request.POST.get("line_user_id"))

                service = None
                if service_id:
                    service = Service.objects.filter(pk=service_id).first()
                    if not service:
                        errors.append("Selected service does not exist.")

                cost_center = None
                if cost_center_id:
                    cost_center = CostCenter.objects.filter(pk=cost_center_id).first()
                    if not cost_center:
                        errors.append("Selected cost centre does not exist.")

                user = None
                if user_id:
                    user = User.objects.filter(pk=user_id).first()
                    if not user:
                        errors.append("Selected user does not exist.")

                if not description:
                    errors.append("Line description is required.")

                quantity = None
                try:
                    quantity = _parse_decimal(quantity_raw)
                except Exception as e:
                    errors.append(str(e))

                line_amount = None
                if amount_raw:
                    try:
                        line_amount = _parse_decimal(amount_raw)
                    except Exception as e:
                        errors.append(str(e))
                else:
                    errors.append("Line amount is required.")

                if errors:
                    for e in errors:
                        messages.error(request, e)
                    return _redirect_back(include_selected=True)

                line = InvoiceLine.objects.create(
                    invoice=invoice,
                    service=service,
                    description=description,
                    quantity=quantity or 1,
                    unit_price=None,
                    line_amount=line_amount,
                    currency=currency,
                    cost_center=cost_center,
                    user=user,
                )

                _audit_log_event(
                    request=request,
                    object_type="Invoice",
                    object_id=invoice.pk,
                    action="update",  # invoice е променен
                    description=(
                        f"Added line #{line.pk}: '{description}', "
                        f"amount {line_amount} {currency}."
                    ),
                )

                messages.success(request, "Invoice line added.")
                return _redirect_back(include_selected=True)

            # 1e) UPDATE header полета (default)
            errors: list[str] = []
            before = _invoice_snapshot(invoice)

            vendor_id = _as_str(request.POST.get("vendor_id"))
            contract_id = _as_str(request.POST.get("contract_id"))
            invoice_number = _as_str(request.POST.get("invoice_number"))
            invoice_date_raw = _as_str(request.POST.get("invoice_date"))
            currency = _as_str(request.POST.get("currency"))
            total_amount_raw = _as_str(request.POST.get("total_amount"))
            tax_amount_raw = _as_str(request.POST.get("tax_amount"))
            period_start_raw = _as_str(request.POST.get("period_start"))
            period_end_raw = _as_str(request.POST.get("period_end"))
            notes = _as_str(request.POST.get("notes"))

            if not invoice_number:
                errors.append("Invoice number is required.")

            vendor = None
            if not vendor_id:
                errors.append("Vendor is required.")
            else:
                vendor = Vendor.objects.filter(pk=vendor_id).first()
                if not vendor:
                    errors.append("Selected vendor does not exist.")

            contract = None
            if contract_id:
                contract = (
                    Contract.objects.filter(owner=request.user, pk=contract_id)
                    .select_related("vendor")
                    .first()
                )
                if not contract:
                    errors.append("Selected contract does not exist.")

            invoice_date = None
            if invoice_date_raw:
                try:
                    invoice_date = _parse_date(invoice_date_raw)
                except Exception as e:
                    errors.append(str(e))

            total_amount = None
            if total_amount_raw:
                try:
                    total_amount = _parse_decimal(total_amount_raw)
                except Exception as e:
                    errors.append(str(e))

            tax_amount = None
            if tax_amount_raw:
                try:
                    tax_amount = _parse_decimal(tax_amount_raw)
                except Exception as e:
                    errors.append(str(e))

            period_start = None
            period_end = None
            try:
                if period_start_raw:
                    period_start = _parse_date(period_start_raw)
                if period_end_raw:
                    period_end = _parse_date(period_end_raw)
            except Exception as e:
                errors.append(str(e))

            if errors:
                for e in errors:
                    messages.error(request, e)
                return _redirect_back(include_selected=True)

            invoice.vendor = vendor
            invoice.contract = contract
            invoice.invoice_number = invoice_number
            if invoice_date:
                invoice.invoice_date = invoice_date
            invoice.currency = currency or invoice.currency
            if total_amount is not None:
                invoice.total_amount = total_amount
            invoice.tax_amount = tax_amount
            invoice.period_start = period_start
            invoice.period_end = period_end
            invoice.notes = notes

            upload_file = request.FILES.get("file")
            if upload_file:
                invoice.file = upload_file

            invoice.save()

            after = _invoice_snapshot(invoice)
            changes = _diff_snapshots(before, after)

            _audit_log_event(
                request=request,
                object_type="Invoice",
                object_id=invoice.pk,
                action="update",
                description="; ".join(changes)
                if changes
                else "Invoice updated (inline).",
            )

            messages.success(request, "Invoice updated successfully.")
            return _redirect_back(include_selected=True)

        # 2) ADD (modal) – няма selected
        else:
            form = InvoiceUploadForm(request.POST, request.FILES)
            if form.is_valid():
                invoice = form.save(owner=request.user)

                _audit_log_event(
                    request=request,
                    object_type="Invoice",
                    object_id=invoice.pk,
                    action="create",
                    description=(
                        f"Created invoice '{invoice.invoice_number}' "
                        f"for vendor '{invoice.vendor}'."
                    ),
                )

                messages.success(
                    request,
                    f"Invoice '{invoice.invoice_number}' saved for vendor {invoice.vendor.name}.",
                )
                if invoice.contract is None:
                    messages.warning(
                        request,
                        "Invoice saved, but no matching contract was linked.",
                    )
                return redirect("portal:invoices")
            else:
                add_form_has_errors = True
    else:
        form = InvoiceUploadForm()

    # -------------------------
    # GET (и POST с грешен Add form)
    # -------------------------
    base_qs = (
        Invoice.objects.filter(owner=request.user)
        .select_related("vendor", "contract")
        .order_by("-invoice_date", "-id")
    )

    total_invoices = base_qs.count()
    total_amount = base_qs.aggregate(total=Sum("total_amount"))["total"] or 0

    paginator = Paginator(base_qs, rows_per_page)
    page_obj = paginator.get_page(page_number)
    invoices_page = list(page_obj.object_list)

    # Selected invoice + breakdown + audit + lines
    selected_invoice = None
    allocation_by_cost_center = []
    service_breakdown = []
    audit_events = []
    invoice_lines = []
    lines_total = None

    if selected_id:
        try:
            selected_invoice = (
                Invoice.objects.filter(owner=request.user)
                .select_related("vendor", "contract")
                .get(pk=int(selected_id))
            )
        except (Invoice.DoesNotExist, ValueError):
            selected_invoice = None

    if selected_invoice:
        lines_qs = selected_invoice.lines.select_related(
            "service", "cost_center", "user", "service__vendor"
        ).order_by("id")

        invoice_lines = list(lines_qs)

        lines_total = (
            lines_qs.aggregate(total=Sum("line_amount"))["total"]
            if lines_qs.exists()
            else None
        )

        allocation_by_cost_center = (
            lines_qs.values("cost_center__code", "cost_center__name")
            .annotate(total=Sum("line_amount"))
            .order_by("cost_center__code")
        )

        service_breakdown = (
            lines_qs.values("service__vendor__name", "service__name")
            .annotate(total=Sum("line_amount"))
            .order_by("service__vendor__name", "service__name")
        )

        audit_events = _audit_fetch_events(
            object_type="Invoice", object_id=selected_invoice.pk, limit=50
        )

    vendors = Vendor.objects.all().order_by("name")
    contracts = (
        Contract.objects.filter(owner=request.user)
        .select_related("vendor")
        .order_by("vendor__name", "contract_name")
    )

    # dropdown-и за Lines & splits
    services = (
        Service.objects.select_related("vendor")
        .all()
        .order_by("vendor__name", "name")
    )
    cost_centers = CostCenter.objects.all().order_by("code")
    users = User.objects.all().order_by("username")

    context = {
        "invoices": invoices_page,
        "total_invoices": total_invoices,
        "total_amount": total_amount,
        "page_obj": page_obj,
        "current_page": page_obj.number,
        "rows_per_page": rows_per_page,
        "rows_options": rows_options,
        "show_closed": show_closed,
        "selected_invoice": selected_invoice,
        "allocation_by_cost_center": allocation_by_cost_center,
        "service_breakdown": service_breakdown,
        "audit_events": audit_events,
        "invoice_lines": invoice_lines,
        "lines_total": lines_total,
        "vendors": vendors,
        "contracts": contracts,
        "services": services,
        "cost_centers": cost_centers,
        "users": users,
        "form": form,
        "add_form_has_errors": add_form_has_errors,
    }
    return render(request, "portal/invoices.html", context)




# ----------
# VENDORS
# ----------

@login_required
def vendor_list(request):
    # -------------------------
    # GET params (Users/Services style)
    # -------------------------
    show_closed = (request.GET.get("show_closed") in ("1", "true", "True", "on", "yes"))

    rows_options = [10, 20, 30, 50, 100, 250]
    try:
        rows_per_page = int(request.GET.get("rows") or 50)
    except (TypeError, ValueError):
        rows_per_page = 50
    if rows_per_page not in rows_options:
        rows_per_page = 50

    page_param = request.GET.get("page") or "1"
    if page_param == "":
        page_param = "1"

    selected_id = _as_str(request.GET.get("selected"))

    # -------------------------
    # Base queryset + show_closed
    # -------------------------
    vendors_qs = (
        Vendor.objects.all()
        .annotate(contract_count=Count("contracts", distinct=True))
        .annotate(invoice_count=Count("invoices", distinct=True))
        .order_by("name")
    )

    if not show_closed and hasattr(Vendor, "is_active"):
        vendors_qs = vendors_qs.filter(is_active=True)

    total_vendors = vendors_qs.count()

    # -------------------------
    # POST:
    #   1) inline update (when selected is present)
    #   2) create vendor (existing form behaviour)
    # -------------------------
    form = VendorCreateForm()
    add_form_has_errors = False

    if request.method == "POST":
        inline_selected = _as_str(request.POST.get("selected"))

        # ---------- INLINE UPDATE ----------
        if inline_selected:
            vendor = get_object_or_404(Vendor, pk=int(inline_selected))

            errors: list[str] = []
            before = _vendor_snapshot(vendor)

            name = _as_str(request.POST.get("name"))
            vendor_type = _as_str(request.POST.get("vendor_type"))
            primary_contact_name = _as_str(request.POST.get("primary_contact_name"))
            primary_contact_email = _as_str(request.POST.get("primary_contact_email"))
            website = _as_str(request.POST.get("website"))
            tags = _as_str(request.POST.get("tags"))
            notes = _as_str(request.POST.get("notes"))

            raw_is_active = _as_str(request.POST.get("is_active"))
            is_active_new = None
            if hasattr(vendor, "is_active"):
                if raw_is_active in ("0", "false", "False", "off", "no"):
                    is_active_new = False
                else:
                    is_active_new = True

            if not name:
                errors.append("Vendor name is required.")

            valid_types = {choice[0] for choice in Vendor.VENDOR_TYPE_CHOICES}
            if vendor_type and vendor_type not in valid_types:
                errors.append("Invalid vendor type.")

            if errors:
                for e in errors:
                    messages.error(request, e)
            else:
                vendor.name = name
                vendor.vendor_type = vendor_type
                vendor.primary_contact_name = primary_contact_name
                vendor.primary_contact_email = primary_contact_email
                vendor.website = website
                vendor.tags = tags
                vendor.notes = notes
                if is_active_new is not None:
                    vendor.is_active = is_active_new
                vendor.save()

                after = _vendor_snapshot(vendor)
                changes = _diff_snapshots(before=before, after=after)
                _audit_log_event(
                    request=request,
                    object_type="Vendor",
                    object_id=vendor.pk,
                    action="update",
                    description="; ".join(changes) if changes else "Vendor updated.",
                )

                if hasattr(vendor, "is_active") and vendor.is_active is False:
                    messages.success(request, "Vendor updated and marked as Closed.")
                else:
                    messages.success(request, "Vendor updated successfully.")

            # back to same selection + keep paging params
            return redirect(
                f"{request.path}?page={_as_str(request.POST.get('page') or '1')}"
                f"&rows={_as_str(request.POST.get('rows') or rows_per_page)}"
                f"&show_closed={_as_str(request.POST.get('show_closed') or ('1' if show_closed else '0'))}"
                f"&selected={vendor.pk}#vendor-details"
            )

        # ---------- CREATE VENDOR (existing behaviour) ----------
        form = VendorCreateForm(request.POST)
        if form.is_valid():
            vendor = form.save()
            messages.success(request, f"Vendor '{vendor.name}' created successfully.")

            _audit_log_event(
                request=request,
                object_type="Vendor",
                object_id=vendor.pk,
                action="create",
                description=f"Created vendor '{vendor.name}'.",
            )

            return redirect("portal:vendors")

        add_form_has_errors = True
        # (form errors will render in template)

    # -------------------------
    # Pagination
    # -------------------------
    paginator = Paginator(vendors_qs, rows_per_page)
    page_obj = paginator.get_page(page_param)
    vendors = list(page_obj.object_list)

    # -------------------------
    # Selected vendor (for inline details) + AUDIT events
    # -------------------------
    selected_vendor = None
    audit_events = []

    if selected_id:
        try:
            selected_vendor = (
                Vendor.objects.annotate(contract_count=Count("contracts", distinct=True))
                .annotate(invoice_count=Count("invoices", distinct=True))
                .filter(pk=int(selected_id))
                .first()
            )
        except (TypeError, ValueError):
            selected_vendor = None

    if selected_vendor:
        audit_events = _audit_fetch_events(object_type="Vendor", object_id=selected_vendor.pk, limit=50)

    context = {
        "vendors": vendors,
        "form": form,
        "add_form_has_errors": add_form_has_errors,

        "show_closed": show_closed,
        "rows_per_page": rows_per_page,
        "rows_options": rows_options,

        "page_obj": page_obj,
        "current_page": page_obj.number,
        "total_vendors": total_vendors,

        "selected_vendor": selected_vendor,
        "vendor_type_choices": Vendor.VENDOR_TYPE_CHOICES,

        # IMPORTANT: used by vendors.html collapse
        "audit_events": audit_events,
    }
    return render(request, "portal/vendors.html", context)


@login_required
def vendor_detail(request, pk):
    """
    Keep this for backwards compatibility / direct links.
    If you want everything ONLY in vendors.html, you can remove this
    route from urls.py later, but leaving it does not hurt.
    """
    vendor = get_object_or_404(Vendor, pk=pk)

    contracts = (
        Contract.objects.filter(owner=request.user, vendor=vendor)
        .select_related("owning_cost_center")
        .order_by("-start_date", "-created_at")
    )

    invoices = (
        Invoice.objects.filter(owner=request.user, vendor=vendor)
        .select_related("contract")
        .order_by("-invoice_date", "-id")
    )

    services = Service.objects.filter(vendor=vendor).order_by("name")

    total_contract_value = contracts.aggregate(total=Sum("annual_value"))["total"] or 0
    total_invoiced = invoices.aggregate(total=Sum("total_amount"))["total"] or 0

    if request.method == "POST":
        action = _as_str(request.POST.get("action")) or "update"

        if action == "delete":
            name = vendor.name
            vendor_id = vendor.pk

            _audit_log_event(
                request=request,
                object_type="Vendor",
                object_id=vendor_id,
                action="delete",
                description=f"Deleted vendor '{name}'.",
            )

            try:
                vendor.delete()
                messages.success(request, f"Vendor '{name}' was deleted.")
                return redirect("portal:vendors")
            except ProtectedError:
                messages.error(
                    request,
                    "This vendor cannot be deleted because there are related contracts or invoices.",
                )

        elif action == "update_status":
            before = _vendor_snapshot(vendor)

            raw = _as_str(request.POST.get("is_active"))
            if raw == "":
                is_active_new = True
            elif raw in ("0", "false", "False", "off", "no"):
                is_active_new = False
            else:
                is_active_new = True

            if hasattr(vendor, "is_active"):
                vendor.is_active = is_active_new
                vendor.save(update_fields=["is_active"])

                after = _vendor_snapshot(vendor)
                changes = _diff_snapshots(before, after)
                _audit_log_event(
                    request=request,
                    object_type="Vendor",
                    object_id=vendor.pk,
                    action="update",
                    description="; ".join(changes) if changes else "Vendor status updated.",
                )

                if is_active_new:
                    messages.success(request, "Vendor marked as Active.")
                    return redirect("portal:vendor_detail", pk=vendor.pk)

                messages.success(request, "Vendor marked as Closed.")
                return redirect("portal:vendors")

            messages.error(request, "Vendor status field is not available yet (missing is_active).")
            return redirect("portal:vendor_detail", pk=vendor.pk)

        else:
            errors: list[str] = []
            before = _vendor_snapshot(vendor)

            name = _as_str(request.POST.get("name"))
            vendor_type = _as_str(request.POST.get("vendor_type"))
            primary_contact_name = _as_str(request.POST.get("primary_contact_name"))
            primary_contact_email = _as_str(request.POST.get("primary_contact_email"))
            website = _as_str(request.POST.get("website"))
            tags = _as_str(request.POST.get("tags"))
            notes = _as_str(request.POST.get("notes"))

            raw_is_active = _as_str(request.POST.get("is_active"))
            is_active_new = None
            if hasattr(vendor, "is_active"):
                if raw_is_active == "":
                    is_active_new = True
                elif raw_is_active in ("0", "false", "False", "off", "no"):
                    is_active_new = False
                else:
                    is_active_new = True

            if not name:
                errors.append("Vendor name is required.")

            valid_types = {choice[0] for choice in Vendor.VENDOR_TYPE_CHOICES}
            if vendor_type and vendor_type not in valid_types:
                errors.append("Invalid vendor type.")

            if errors:
                for e in errors:
                    messages.error(request, e)
            else:
                vendor.name = name
                vendor.vendor_type = vendor_type
                vendor.primary_contact_name = primary_contact_name
                vendor.primary_contact_email = primary_contact_email
                vendor.website = website
                vendor.tags = tags
                vendor.notes = notes
                if is_active_new is not None:
                    vendor.is_active = is_active_new
                vendor.save()

                after = _vendor_snapshot(vendor)
                changes = _diff_snapshots(after=after, before=before)
                _audit_log_event(
                    request=request,
                    object_type="Vendor",
                    object_id=vendor.pk,
                    action="update",
                    description="; ".join(changes) if changes else "Vendor updated.",
                )

                if hasattr(vendor, "is_active") and vendor.is_active is False:
                    messages.success(request, "Vendor updated and marked as Closed.")
                    return redirect("portal:vendors")

                messages.success(request, "Vendor updated successfully.")
                return redirect("portal:vendor_detail", pk=vendor.pk)

    audit_events = _audit_fetch_events(object_type="Vendor", object_id=vendor.pk, limit=50)

    context = {
        "vendor": vendor,
        "contracts": contracts,
        "invoices": invoices,
        "services": services,
        "total_contract_value": total_contract_value,
        "total_invoiced": total_invoiced,
        "vendor_type_choices": Vendor.VENDOR_TYPE_CHOICES,
        "audit_events": audit_events,
    }
    return render(request, "portal/vendor_detail.html", context)


# ----------
# SERVICES
# ----------

@login_required
def service_list(request):
    vendors = Vendor.objects.all().order_by("name")

    # -------------------------
    # GET params (Users-style)
    # -------------------------
    show_closed = (request.GET.get("show_closed") in ("1", "true", "True", "on", "yes"))

    rows_options = [10, 20, 30, 50, 100, 250]
    try:
        rows_per_page = int(request.GET.get("rows") or 50)
    except (TypeError, ValueError):
        rows_per_page = 50
    if rows_per_page not in rows_options:
        rows_per_page = 50

    page_param = request.GET.get("page") or "1"
    if page_param == "":
        page_param = "1"

    selected_id = _as_str(request.GET.get("selected"))

    # -------------------------
    # Base queryset
    # -------------------------
    services_qs = (
        Service.objects.select_related("vendor", "primary_contract")
        .order_by("vendor__name", "name")
    )
    if not show_closed and hasattr(Service, "is_active"):
        services_qs = services_qs.filter(is_active=True)

    total_services = services_qs.count()

    # -------------------------
    # POST: inline update OR add modal create
    # -------------------------
    add_form_data: dict = {}
    add_form_has_errors = False

    if request.method == "POST":
        inline_selected = _as_str(request.POST.get("selected"))

        # ---------- INLINE UPDATE (when selected is present) ----------
        if inline_selected:
            service = get_object_or_404(
                Service.objects.select_related("vendor", "primary_contract"),
                pk=int(inline_selected),
            )

            errors: list[str] = []
            before = _service_snapshot(service)

            vendor_id = _as_str(request.POST.get("vendor_id"))
            name = _as_str(request.POST.get("name"))
            category = _as_str(request.POST.get("category"))
            billing_frequency = _as_str(request.POST.get("billing_frequency"))
            default_currency = _as_str(request.POST.get("default_currency"))
            service_code = _as_str(request.POST.get("service_code"))

            # allow both names (old + new)
            owner_display = _as_str(
                request.POST.get("service_owner")
                or request.POST.get("owner_display")
            )

            allocation_split = _as_str(request.POST.get("allocation_split"))
            list_price_raw = _as_str(request.POST.get("list_price"))

            # allow both: old contract_ref (text) and new primary_contract_id (id)
            contract_ref = _as_str(request.POST.get("contract_ref"))
            primary_contract_id = _as_str(request.POST.get("primary_contract_id"))

            # status (optional field on model)
            is_active_new = None
            if hasattr(service, "is_active"):
                raw_is_active = _as_str(request.POST.get("is_active"))
                if raw_is_active in ("0", "false", "False", "off", "no"):
                    is_active_new = False
                else:
                    is_active_new = True

            # validate vendor/name
            vendor = None
            if not vendor_id:
                errors.append("Vendor is required.")
            else:
                vendor = Vendor.objects.filter(pk=vendor_id).first()
                if not vendor:
                    errors.append("Selected vendor does not exist.")

            if not name:
                errors.append("Service name is required.")

            # parse list_price
            list_price = None
            if list_price_raw:
                try:
                    list_price = _parse_decimal(list_price_raw)
                except Exception as e:
                    errors.append(str(e))

            # primary contract resolve:
            # 1) if primary_contract_id present -> exact id
            # 2) else if contract_ref present -> match by name/id/pk (old logic)
            primary_contract = None
            contract_not_found = False

            if primary_contract_id:
                pc = Contract.objects.filter(owner=request.user, pk=primary_contract_id).first()
                if pc:
                    primary_contract = pc
                else:
                    contract_not_found = True
            elif contract_ref and vendor:
                contract_filters = Q(contract_name__iexact=contract_ref) | Q(contract_id__iexact=contract_ref)
                try:
                    ref_pk = int(contract_ref)
                    contract_filters |= Q(pk=ref_pk)
                except (TypeError, ValueError):
                    pass

                primary_contract = (
                    Contract.objects.filter(owner=request.user, vendor=vendor)
                    .filter(contract_filters)
                    .first()
                )
                if primary_contract is None:
                    contract_not_found = True

            # uniqueness check
            if vendor and name:
                exists = (
                    Service.objects.filter(vendor=vendor, name__iexact=name)
                    .exclude(pk=service.pk)
                    .exists()
                )
                if exists:
                    errors.append("A service with this name already exists for the selected vendor.")

            # redirect helper (keep state)
            post_page = _as_str(request.POST.get("page") or "1") or "1"
            post_rows = _as_str(request.POST.get("rows") or rows_per_page)
            post_show_closed = _as_str(
                request.POST.get("show_closed") or ("1" if show_closed else "0")
            )

            if errors:
                for e in errors:
                    messages.error(request, e)

                return redirect(
                    f"{request.path}?page={post_page}"
                    f"&rows={post_rows}"
                    f"&show_closed={post_show_closed}"
                    f"&selected={service.pk}#service-details"
                )

            # save
            service.vendor = vendor
            service.name = name
            service.category = category or ""
            service.default_billing_frequency = billing_frequency or ""
            service.default_currency = default_currency or ""
            service.service_code = service_code or ""
            service.owner_display = owner_display or ""
            service.allocation_split = allocation_split or ""
            service.list_price = list_price
            service.primary_contract = primary_contract
            if is_active_new is not None:
                service.is_active = is_active_new
            service.save()

            after = _service_snapshot(service)
            changes = _diff_snapshots(before, after)
            _audit_log_event(
                request=request,
                object_type="Service",
                object_id=service.pk,
                action="update",
                description="; ".join(changes) if changes else "Service updated.",
            )

            messages.success(request, "Service updated successfully.")
            if contract_not_found and (contract_ref or primary_contract_id):
                messages.warning(request, "Service saved, but no matching contract was linked.")

            return redirect(
                f"{request.path}?page={post_page}"
                f"&rows={post_rows}"
                f"&show_closed={post_show_closed}"
                f"&selected={service.pk}#service-details"
            )

        # ---------- ADD MODAL CREATE (existing behavior) ----------
        vendor_id = _as_str(request.POST.get("vendor_id"))
        name = _as_str(request.POST.get("name") or request.POST.get("service_name"))
        category = _as_str(request.POST.get("category"))
        billing_frequency = _as_str(request.POST.get("billing_frequency"))
        default_currency = _as_str(request.POST.get("default_currency"))
        service_code = _as_str(request.POST.get("service_code"))
        owner_display = _as_str(request.POST.get("service_owner") or request.POST.get("owner_display"))
        allocation_split = _as_str(request.POST.get("allocation_split"))
        list_price_raw = _as_str(request.POST.get("list_price"))
        contract_ref = _as_str(request.POST.get("contract_ref"))

        add_form_data = {
            "vendor_id": vendor_id,
            "name": name,
            "category": category,
            "billing_frequency": billing_frequency,
            "default_currency": default_currency,
            "service_code": service_code,
            "owner_display": owner_display,
            "allocation_split": allocation_split,
            "list_price": list_price_raw,
            "contract_ref": contract_ref,
        }

        errors: list[str] = []

        vendor = None
        if not vendor_id:
            errors.append("Vendor is required.")
        else:
            vendor = Vendor.objects.filter(pk=vendor_id).first()
            if not vendor:
                errors.append("Selected vendor does not exist.")

        if not name:
            errors.append("Service name is required.")

        list_price = None
        if list_price_raw:
            try:
                list_price = _parse_decimal(list_price_raw)
            except Exception as e:
                errors.append(str(e))

        primary_contract = None
        contract_not_found = False
        if contract_ref and vendor:
            contract_filters = Q(contract_name__iexact=contract_ref) | Q(contract_id__iexact=contract_ref)
            try:
                ref_pk = int(contract_ref)
                contract_filters |= Q(pk=ref_pk)
            except (TypeError, ValueError):
                pass

            primary_contract = (
                Contract.objects.filter(owner=request.user, vendor=vendor)
                .filter(contract_filters)
                .first()
            )
            if primary_contract is None:
                contract_not_found = True

        if vendor and name:
            exists = Service.objects.filter(vendor=vendor, name__iexact=name).exists()
            if exists:
                errors.append("A service with this name already exists for the selected vendor.")

        if errors:
            add_form_has_errors = True
            for e in errors:
                messages.error(request, e)
        else:
            service = Service.objects.create(
                vendor=vendor,
                name=name,
                category=category or "",
                service_code=service_code or "",
                default_currency=default_currency or "",
                default_billing_frequency=billing_frequency or "",
                owner_display=owner_display or "",
                allocation_split=allocation_split or "",
                list_price=list_price,
                primary_contract=primary_contract,
            )

            _audit_log_event(
                request=request,
                object_type="Service",
                object_id=service.pk,
                action="create",
                description=f"Created service '{service}'.",
            )

            messages.success(request, "Service created successfully.")
            if contract_not_found:
                messages.warning(request, "Service saved, but no matching contract was linked.")
            return redirect("portal:services")

    # -------------------------
    # Pagination
    # -------------------------
    paginator = Paginator(services_qs, rows_per_page)
    page_obj = paginator.get_page(page_param)
    services = list(page_obj.object_list)

    # Selected service (for inline details)
    selected_service = None
    if selected_id:
        try:
            selected_service = (
                Service.objects.select_related("vendor", "primary_contract")
                .filter(pk=int(selected_id))
                .first()
            )
        except (TypeError, ValueError):
            selected_service = None

    # ---------- Audit events for selected service (for inline Audit tab) ----------
    audit_events = []
    if selected_service:
        audit_events = _audit_fetch_events(
            object_type="Service",
            object_id=selected_service.pk,
            limit=50,
        )

    # IMPORTANT: template uses current_page
    try:
        current_page = int(page_obj.number)
    except Exception:
        current_page = 1

    context = {
        "services": services,
        "vendors": vendors,

        "show_closed": show_closed,
        "rows_per_page": rows_per_page,
        "rows_options": rows_options,

        "page_obj": page_obj,
        "current_page": current_page,
        "total_services": total_services,

        "selected_service": selected_service,
        "audit_events": audit_events,          # <--- новото

        # add modal state
        "add_form_data": add_form_data,
        "add_form_has_errors": add_form_has_errors,
    }
    return render(request, "portal/services.html", context)

@login_required
def service_detail(request, pk):
    service = get_object_or_404(
        Service.objects.select_related("vendor", "primary_contract"),
        pk=pk,
    )

    # Ако имаш нещо по-специално за service detail, може да го допълним после
    context = {
        "service": service,
    }
    return render(request, "portal/service_detail.html", context)


# ----------
# COST CENTERS
# ----------

@login_required
def cost_centers_list(request):
    cost_centers = (
        CostCenter.objects.all()
        .annotate(contract_count=Count("contracts", distinct=True))
        .annotate(line_count=Count("invoice_lines", distinct=True))
        .order_by("code")
    )
    return render(request, "portal/cost_centers.html", {"cost_centers": cost_centers})



# ----------
# USERS
# ----------

from django.core.paginator import Paginator


@login_required
def users_list(request):
    show_closed = (request.GET.get("show_closed") in ("1", "true", "True", "on", "yes"))

    # rows per page (allowlist)
    try:
        rows_per_page = int(request.GET.get("rows") or 50)
    except (TypeError, ValueError):
        rows_per_page = 50

    allowed_rows = [10, 20, 30, 50]
    if rows_per_page not in allowed_rows:
        rows_per_page = 50

    # page (robust: page="" -> 1)
    raw_page = (request.GET.get("page") or "").strip()
    try:
        page_number = int(raw_page) if raw_page else 1
    except ValueError:
        page_number = 1
    if page_number < 1:
        page_number = 1

    # selected user id (optional)
    raw_selected = (request.GET.get("selected") or "").strip()
    try:
        selected_id = int(raw_selected) if raw_selected else None
    except ValueError:
        selected_id = None

    # base queryset (ensure profiles exist)
    base_qs = User.objects.all().order_by("username")
    if not show_closed:
        base_qs = base_qs.filter(is_active=True)

    # ensure UserProfile exists for each user (keep your behaviour, but safer)
    for u in base_qs:
        UserProfile.objects.get_or_create(user=u)

    # real queryset for screen
    users_qs = (
        User.objects.select_related("profile", "profile__cost_center", "profile__manager")
        .order_by("username")
    )
    if not show_closed:
        users_qs = users_qs.filter(is_active=True)

    total_users = users_qs.count()

    # -------------------------
    # Inline SAVE (POST)
    # -------------------------
    if request.method == "POST":
        # we expect selected user id in POST
        raw_post_selected = (request.POST.get("selected") or "").strip()
        try:
            post_selected_id = int(raw_post_selected)
        except (TypeError, ValueError):
            post_selected_id = None

        if not post_selected_id:
            messages.error(request, "No user selected.")
            return redirect("portal:users")

        user_obj = get_object_or_404(
            User.objects.select_related("profile", "profile__cost_center", "profile__manager"),
            pk=post_selected_id,
        )
        profile, _ = UserProfile.objects.get_or_create(user=user_obj)

        before = _user_snapshot(user_obj, profile)

        # ---- account fields ----
        username = _as_str(request.POST.get("username"))
        email = _as_str(request.POST.get("email"))

        raw_is_active = _as_str(request.POST.get("is_active"))
        is_active_flag = False if raw_is_active in ("0", "false", "False", "off", "no") else True

        # ---- profile fields (safe) ----
        full_name = _as_str(request.POST.get("full_name"))
        location = _as_str(request.POST.get("location"))
        legal_entity = _as_str(request.POST.get("legal_entity"))

        # support both "phone" and "phone_number"
        phone_number = _as_str(request.POST.get("phone_number")) or _as_str(request.POST.get("phone"))

        # cost center: accept either id or code
        cost_center = None
        cost_center_id = _as_str(request.POST.get("cost_center_id"))
        cost_center_code = _as_str(request.POST.get("cost_center"))
        if cost_center_id:
            cost_center = CostCenter.objects.filter(pk=cost_center_id).first()
        elif cost_center_code:
            cost_center = CostCenter.objects.filter(code__iexact=cost_center_code).first()

        # manager: accept either id or username
        manager = None
        manager_id = _as_str(request.POST.get("manager_id"))
        manager_username = _as_str(request.POST.get("manager"))
        if manager_id:
            manager = User.objects.filter(pk=manager_id).first()
        elif manager_username:
            manager = User.objects.filter(username__iexact=manager_username).first()

        errors: list[str] = []

        # validations (mirror your user_detail style)
        if not username:
            errors.append("Username is required.")
        else:
            if User.objects.exclude(pk=user_obj.pk).filter(username__iexact=username).exists():
                errors.append("Another user with this username already exists.")

        if email:
            if User.objects.exclude(pk=user_obj.pk).filter(email__iexact=email).exists():
                errors.append("Another user with this email already exists.")

        if (cost_center_id or cost_center_code) and not cost_center:
            errors.append("Selected cost centre does not exist.")

        if (manager_id or manager_username) and manager_id and not manager:
            errors.append("Selected manager does not exist.")

        if errors:
            for e in errors:
                messages.error(request, e)
            # keep same paging state after error
            return redirect(
                f"{request.path}?page={page_number}&rows={rows_per_page}"
                f"&show_closed={'1' if show_closed else '0'}&selected={user_obj.pk}#user-details"
            )

        # persist
        user_obj.username = username
        user_obj.email = email
        user_obj.is_active = is_active_flag
        user_obj.save()

        profile.full_name = full_name
        profile.cost_center = cost_center
        profile.manager = manager
        profile.location = location
        profile.legal_entity = legal_entity
        # keep compatibility with your model field name
        if hasattr(profile, "phone_number"):
            profile.phone_number = phone_number
        elif hasattr(profile, "phone"):
            profile.phone = phone_number
        profile.save()

        after = _user_snapshot(user_obj, profile)
        changes = _diff_snapshots(before, after)
        _audit_log_event(
            request=request,
            object_type="User",
            object_id=user_obj.pk,
            action="update",
            description="; ".join(changes) if changes else "User updated (inline).",
        )

        messages.success(request, "User updated successfully.")

        # redirect back, keep same page/rows/show_closed/selected
        return redirect(
            f"{request.path}?page={page_number}&rows={rows_per_page}"
            f"&show_closed={'1' if show_closed else '0'}&selected={user_obj.pk}#user-details"
        )

    # -------------------------
    # Pagination
    # -------------------------
    paginator = Paginator(users_qs, rows_per_page)
    page_obj = paginator.get_page(page_number)
    users_page = page_obj.object_list

    # -------------------------
    # Selected user + services (for inline details)
    # -------------------------
    selected_user = None
    selected_services = []
    audit_events = []  # NEW: audit лог за избрания user

    if selected_id:
        selected_user = User.objects.select_related(
            "profile", "profile__cost_center", "profile__manager"
        ).filter(pk=selected_id).first()

        if selected_user:
            UserProfile.objects.get_or_create(user=selected_user)

            assignments = (
                ServiceAssignment.objects
                .filter(user=selected_user)
                .select_related("service", "service__vendor")
                .order_by("service__vendor__name", "service__name")
            )

            for a in assignments:
                svc = a.service
                if not svc:
                    continue

                vendor_name = svc.vendor.name if getattr(svc, "vendor", None) else "—"
                price = getattr(svc, "list_price", None)
                ccy = getattr(svc, "default_currency", "") or "—"
                is_active = getattr(svc, "is_active", True)

                selected_services.append({
                    "service_name": getattr(svc, "name", "—"),
                    "vendor_name": vendor_name,
                    "price": price if price is not None else "—",
                    "ccy": ccy,
                    "status": "Active" if is_active else "Closed",
                })

            # >>> ТУК: взимаме audit log за User, както при Services/Contracts
            audit_events = _audit_fetch_events(
                object_type="User",
                object_id=selected_user.pk,
                limit=50,
            )

    rows_options = [10, 20, 30, 50]

    return render(
        request,
        "portal/users.html",
        {
            "users": users_page,
            "page_obj": page_obj,
            "current_page": page_obj.number,
            "show_closed": show_closed,
            "rows_per_page": rows_per_page,
            "rows_options": rows_options,
            "total_users": total_users,
            "selected_user": selected_user,
            "selected_services": selected_services,
            "audit_events": audit_events,  # NEW: подаваме към шаблона
        },
    )





# ----------
# PERMISSIONS (FIXED: bulk assign/unassign with buttons)
# ----------

@login_required
def permissions(request):
    def _flag(name: str) -> bool:
        v = _as_str(request.POST.get(name) or request.GET.get(name))
        return v in ("1", "true", "True", "on", "yes")

    show_closed_users = _flag("show_closed_users")
    show_closed_services = _flag("show_closed_services")
    show_closed_vendors = _flag("show_closed_vendors")

    vendors = Vendor.objects.all().order_by("name")
    if not show_closed_vendors and hasattr(Vendor, "is_active"):
        vendors = vendors.filter(is_active=True)

    vendor_id = _as_str(request.GET.get("vendor_id") or request.POST.get("vendor_id"))
    selected_vendor = Vendor.objects.filter(pk=vendor_id).first() if vendor_id else None

    users_qs = User.objects.select_related("profile", "profile__cost_center").order_by("username")
    if not show_closed_users:
        users_qs = users_qs.filter(is_active=True)

    for u in users_qs:
        UserProfile.objects.get_or_create(user=u)

    services_qs = Service.objects.none()
    if selected_vendor:
        services_qs = Service.objects.filter(vendor=selected_vendor).order_by("name")
        if not show_closed_services and hasattr(Service, "is_active"):
            services_qs = services_qs.filter(is_active=True)

    if request.method == "POST":
        action = _as_str(request.POST.get("action")) or "assign"
        user_ids = request.POST.getlist("user_ids")
        service_ids = request.POST.getlist("service_ids")

        if not selected_vendor:
            messages.error(request, "Vendor is required.")
        elif not user_ids or not service_ids:
            messages.error(request, "Select at least 1 user and 1 service.")
        else:
            users_sel = User.objects.filter(pk__in=user_ids)
            services_sel = Service.objects.filter(pk__in=service_ids, vendor=selected_vendor)

            if not users_sel.exists():
                messages.error(request, "No valid users selected.")
            elif not services_sel.exists():
                messages.error(request, "No valid services selected for this vendor.")
            else:
                created_count = 0
                deleted_count = 0

                with transaction.atomic():
                    if action == "assign":
                        for u in users_sel:
                            for s in services_sel:
                                obj, created = ServiceAssignment.objects.get_or_create(
                                    user=u,
                                    service=s,
                                    defaults={"assigned_by": request.user},
                                )
                                if created:
                                    created_count += 1
                                    _audit_log_event(
                                        request=request,
                                        object_type="User",
                                        object_id=u.pk,
                                        action="update",
                                        description=f"Assigned service: {s.vendor.name} – {s.name}",
                                    )

                        messages.success(
                            request,
                            f"Assigned {created_count} permission(s) (users: {users_sel.count()}, services: {services_sel.count()})."
                        )

                    elif action == "unassign":
                        qs = ServiceAssignment.objects.filter(user__in=users_sel, service__in=services_sel)
                        pairs = list(qs.select_related("service", "service__vendor", "user"))
                        deleted_count, _ = qs.delete()

                        for p in pairs:
                            _audit_log_event(
                                request=request,
                                object_type="User",
                                object_id=p.user_id,
                                action="update",
                                description=f"Unassigned service: {p.service.vendor.name} – {p.service.name}",
                            )

                        messages.success(
                            request,
                            f"Unassigned {deleted_count} permission(s) (users: {users_sel.count()}, services: {services_sel.count()})."
                        )
                    else:
                        messages.error(request, "Unknown action.")

        # preserve vendor + toggles on redirect
        qs = ""
        if selected_vendor:
            qs = f"vendor_id={selected_vendor.id}"
        else:
            qs = ""

        if show_closed_users:
            qs += ("&" if qs else "") + "show_closed_users=1"
        if show_closed_services:
            qs += ("&" if qs else "") + "show_closed_services=1"
        if show_closed_vendors:
            qs += ("&" if qs else "") + "show_closed_vendors=1"

        return redirect(f"{reverse('portal:permissions')}?{qs}" if qs else reverse("portal:permissions"))

    return render(request, "portal/permissions.html", {
        "vendors": vendors,
        "selected_vendor": selected_vendor,
        "users": users_qs,
        "services": services_qs,
        "show_closed_users": show_closed_users,
        "show_closed_services": show_closed_services,
        "show_closed_vendors": show_closed_vendors,
    })


@login_required
def permissions_toggle(request):
    """
    Backward-compatible endpoint (if portal/urls.py still references it).
    Not used by the bulk permissions UI, but prevents AttributeError and can support AJAX toggles.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required."}, status=405)

    user_id = _as_str(request.POST.get("user_id"))
    service_id = _as_str(request.POST.get("service_id"))
    assigned = _as_str(request.POST.get("assigned"))

    if not user_id or not service_id:
        return JsonResponse({"ok": False, "error": "Missing user_id/service_id."}, status=400)

    u = User.objects.filter(pk=user_id).first()
    s = Service.objects.filter(pk=service_id).select_related("vendor").first()
    if not u or not s:
        return JsonResponse({"ok": False, "error": "User or Service not found."}, status=404)

    want_assigned = assigned in ("1", "true", "True", "on", "yes")

    if want_assigned:
        obj, created = ServiceAssignment.objects.get_or_create(
            user=u,
            service=s,
            defaults={"assigned_by": request.user},
        )
        if created:
            _audit_log_event(
                request=request,
                object_type="User",
                object_id=u.pk,
                action="update",
                description=f"Assigned service: {s.vendor.name} – {s.name}",
            )
        return JsonResponse({"ok": True, "assigned": True})

    deleted, _ = ServiceAssignment.objects.filter(user=u, service=s).delete()
    if deleted:
        _audit_log_event(
            request=request,
            object_type="User",
            object_id=u.pk,
            action="update",
            description=f"Unassigned service: {s.vendor.name} – {s.name}",
        )
    return JsonResponse({"ok": True, "assigned": False})


# ----------
# SEARCH (global)
# ----------

@login_required
def global_search(request):
    query = _as_str(request.GET.get("q"))

    vendors = []
    services = []
    contracts = []
    invoices = []
    users = []

    if query:
        vendors = (
            Vendor.objects.filter(
                Q(name__icontains=query)
                | Q(tags__icontains=query)
                | Q(primary_contact_name__icontains=query)
                | Q(primary_contact_email__icontains=query)
            )
            .order_by("name")[:25]
        )

        services = (
            Service.objects.select_related("vendor")
            .filter(
                Q(name__icontains=query)
                | Q(category__icontains=query)
                | Q(service_code__icontains=query)
                | Q(vendor__name__icontains=query)
            )
            .order_by("vendor__name", "name")[:25]
        )

        contracts = (
            Contract.objects.select_related("vendor", "owning_cost_center")
            .filter(owner=request.user)
            .filter(
                Q(contract_name__icontains=query)
                | Q(contract_id__icontains=query)
                | Q(vendor__name__icontains=query)
            )
            .order_by("-start_date", "-created_at")[:25]
        )

        invoices = (
            Invoice.objects.select_related("vendor", "contract")
            .filter(owner=request.user)
            .filter(
                Q(invoice_number__icontains=query)
                | Q(vendor__name__icontains=query)
                | Q(contract__contract_name__icontains=query)
            )
            .order_by("-invoice_date", "-id")[:25]
        )

        users = (
            User.objects.select_related("profile", "profile__cost_center")
            .filter(
                Q(username__icontains=query)
                | Q(first_name__icontains=query)
                | Q(last_name__icontains=query)
                | Q(email__icontains=query)
                | Q(profile__full_name__icontains=query)
            )
            .order_by("username")[:25]
        )

    has_results = bool(query and (vendors or services or contracts or invoices or users))

    return render(
        request,
        "portal/search_results.html",
        {
            "query": query,
            "vendors": vendors,
            "services": services,
            "contracts": contracts,
            "invoices": invoices,
            "users": users,
            "has_results": has_results,
        },
    )


# ----------
# DATA HUB
# ----------

@login_required
def data_hub(request):
    items = []
    for key, cfg in DATA_ENTITIES.items():
        if key == "vendors":
            count = Vendor.objects.count()
        elif key == "cost-centers":
            count = CostCenter.objects.count()
        elif key == "services":
            count = Service.objects.count()
        elif key == "contracts":
            count = Contract.objects.filter(owner=request.user).count()
        elif key == "invoices":
            count = Invoice.objects.filter(owner=request.user).count()
        elif key == "users":
            count = User.objects.count()
        elif key == "permissions":
            count = ServiceAssignment.objects.count()
        else:
            count = 0

        items.append({"key": key, "label": cfg["label"], "count": count})

    return render(request, "portal/data_hub.html", {"items": items})


@login_required
def data_import(request, entity: str):
    cfg = _get_entity_or_404(entity)

    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Please choose a CSV or XLSX file to upload.")
            return redirect("portal:data_import", entity=entity)

        fmt = _detect_format(request, upload.name)

        try:
            rows = _read_table(upload, fmt)
            if not rows:
                messages.warning(request, "The uploaded file has no data rows.")
                return redirect("portal:data_import", entity=entity)

            result = cfg["importer"](rows, request.user)
            messages.success(
                request,
                f"{cfg['label']}: import completed. Created: {result.get('created', 0)}, updated: {result.get('updated', 0)}."
            )
            return redirect("portal:data_hub")

        except Exception as e:
            messages.error(request, f"{cfg['label']}: import failed. {e}")
            return redirect("portal:data_import", entity=entity)

    return render(
        request,
        "portal/data_import.html",
        {
            "entity": entity,
            "label": cfg["label"],
            "template_headers": cfg["template_headers"],
        },
    )


@login_required
def data_export(request, entity: str):
    cfg = _get_entity_or_404(entity)
    fmt = _detect_format(request)

    headers = cfg["template_headers"]
    rows = cfg["exporter"](request.user)

    filename_base = f"datanaut_{entity}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    if fmt == "xlsx":
        return _workbook_response(f"{filename_base}.xlsx", headers, rows)
    return _csv_response(f"{filename_base}.csv", headers, rows)


@login_required
def data_template(request, entity: str):
    cfg = _get_entity_or_404(entity)
    fmt = _detect_format(request)

    headers = cfg["template_headers"]
    rows: list[list[str]] = []

    filename_base = f"template_{entity}"
    if fmt == "xlsx":
        return _workbook_response(f"{filename_base}.xlsx", headers, rows)
    return _csv_response(f"{filename_base}.csv", headers, rows)


# ----------
# PROVISIONING HUB
# ----------

from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.contrib.auth.models import Group
from django.views.decorators.http import require_POST

from .models import (
    ProvisioningRequest,
    Service,
    ServiceAssignment,
    UserProfile,
    Vendor,
)

User = get_user_model()

PROV_ACTING_SESSION_KEY = "prov_acting_user_id"


def _is_prov_admin(user) -> bool:
    return (
        getattr(user, "is_superuser", False)
        or getattr(user, "is_staff", False)
        or user.groups.filter(name__in=["Provisioning Hub Admins", "ProvisioningHubAdmins"]).exists()
    )


def _can_act_for(manager_user, target_user) -> bool:
    """
    Who can manage access for whom:
      - prov admin: can act for anyone
      - user can act for self
      - line manager can act for direct reports (UserProfile.manager == manager_user)
    """
    if not manager_user.is_authenticated:
        return False

    if manager_user.pk == target_user.pk:
        return True

    if _is_prov_admin(manager_user):
        return True

    # Manager rule (direct reports only)
    try:
        profile = getattr(target_user, "profile", None)
        if profile and profile.manager_id == manager_user.pk:
            return True
    except Exception:
        pass

    return False

def is_portal_admin(user) -> bool:
    """
    Кой вижда целия „админ“ портал:
      - superuser
      - staff
      - член на група PortalAdmins
    """
    if not user.is_authenticated:
        return False
    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        return True
    return user.groups.filter(name="PortalAdmins").exists()

def _get_acting_user(request):
    """
    Returns the effective user for provisioning actions.
    If a session acting user exists but is not allowed anymore -> clear it.
    """
    acting_id = request.session.get(PROV_ACTING_SESSION_KEY)
    if not acting_id:
        return request.user

    try:
        target = User.objects.select_related("profile").get(pk=int(acting_id))
    except Exception:
        request.session.pop(PROV_ACTING_SESSION_KEY, None)
        return request.user

    if not _can_act_for(request.user, target):
        request.session.pop(PROV_ACTING_SESSION_KEY, None)
        return request.user

    return target


def _get_manageable_users(request):
    """
    List of users visible in the "Manage access for" dropdown.
    - prov admin: all users (safe for demo; later you can scope by tenant)
    - manager: direct reports only
    """
    if _is_prov_admin(request.user):
        return User.objects.order_by("username")

    # direct reports
    return (
        User.objects.filter(profile__manager=request.user)
        .order_by("username")
    )


@require_POST
@login_required
def provisioning_acting_set(request):
    """
    Sets the acting user (manage access for...) in session.
    POST: user_id, optional next
    """
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "portal:provisioning_hub"
    user_id = request.POST.get("user_id")

    if not user_id or not str(user_id).isdigit():
        messages.error(request, "Invalid selection.")
        return redirect(next_url)

    target = get_object_or_404(User, pk=int(user_id))

    if not _can_act_for(request.user, target):
        messages.error(request, "You do not have permission to manage access for this user.")
        return redirect(next_url)

    request.session[PROV_ACTING_SESSION_KEY] = target.pk
    messages.info(request, f"Managing access for: {target.username}.")
    return redirect(next_url)


@require_POST
@login_required
def provisioning_acting_clear(request):
    """
    Clears acting user (back to self).
    """
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "portal:provisioning_hub"
    request.session.pop(PROV_ACTING_SESSION_KEY, None)
    messages.info(request, "Managing access for: you.")
    return redirect(next_url)


@login_required
@login_required
def provisioning_hub(request):
    """
    Provisioning Hub landing page:
      - summary tiles (counts + last request)
      - left: current assigned services (from ServiceAssignment)
      - right: user card (profile + total cost)
    IMPORTANT: uses acting user if set.
    """
    is_prov_admin = _is_prov_admin(request.user)

    acting_user = _get_acting_user(request)
    profile, _ = UserProfile.objects.get_or_create(user=acting_user)

    assignments = (
        ServiceAssignment.objects
        .filter(user=acting_user)
        .select_related("service", "service__vendor")
        .order_by("service__vendor__name", "service__name")
    )

    assigned_rows = []
    total_cost = Decimal("0")
    currencies = set()

    for a in assignments:
        s = a.service
        if not s:
            continue

        vendor_name = s.vendor.name if s.vendor else "—"
        is_active = getattr(s, "is_active", True)
        list_price = getattr(s, "list_price", None)
        currency = getattr(s, "default_currency", "") or "—"

        assigned_rows.append({
            "service": s,
            "vendor_name": vendor_name,
            "is_active": is_active,
            "list_price": list_price,
            "currency": currency,
        })

        if list_price is not None:
            try:
                total_cost += Decimal(str(list_price))
            except Exception:
                pass

        if currency and currency != "—":
            currencies.add(currency)

    services_count = len(assigned_rows)

    pending_requests_count = (
        ProvisioningRequest.objects.filter(
            requester=acting_user,
            status=ProvisioningRequest.STATUS_PENDING,
        ).count()
    )

    last_request = (
        ProvisioningRequest.objects
        .select_related("service", "service__vendor")
        .filter(requester=acting_user)
        .order_by("-created_at")
        .first()
    )

    approvals_open = 0
    if is_prov_admin:
        approvals_open = ProvisioningRequest.objects.filter(
            status=ProvisioningRequest.STATUS_PENDING
        ).count()

    primary_currency = list(currencies)[0] if len(currencies) == 1 else None

    return render(
        request,
        "portal/provisioning_hub.html",
        {
            "is_prov_admin": is_prov_admin,
            "acting_user": acting_user,
            "is_acting": (acting_user.pk != request.user.pk),
            "manageable_users": _get_manageable_users(request),
            "profile": profile,
            "assigned_rows": assigned_rows,
            "total_cost": total_cost,
            "services_count": services_count,
            "pending_requests_count": pending_requests_count,
            "last_request": last_request,
            "approvals_open": approvals_open,
            "primary_currency": primary_currency,
        },
    )



@login_required
def provisioning_catalog(request):
    """
    Catalog page (active services), rendered with checkboxes.
    Uses acting user for assigned/pending checks.
    """
    acting_user = _get_acting_user(request)

    services = Service.objects.select_related("vendor").order_by("vendor__name", "name")

    if hasattr(Service, "is_active"):
        services = services.filter(is_active=True)

    if hasattr(Vendor, "is_active"):
        services = services.filter(vendor__is_active=True)

    assigned_service_ids = set(
        ServiceAssignment.objects.filter(user=acting_user).values_list("service_id", flat=True)
    )

    pending_service_ids = set(
        ProvisioningRequest.objects.filter(
            requester=acting_user,
            status=ProvisioningRequest.STATUS_PENDING,
        ).values_list("service_id", flat=True)
    )

    by_vendor: dict[str, list] = {}
    for s in services:
        vname = s.vendor.name if s.vendor else "—"
        by_vendor.setdefault(vname, []).append({
            "service": s,
            "is_assigned": (s.id in assigned_service_ids),
            "is_pending": (s.id in pending_service_ids),
        })

    return render(
        request,
        "portal/provisioning_catalog.html",
        {
            "acting_user": acting_user,
            "is_acting": (acting_user.pk != request.user.pk),
            "services_by_vendor": by_vendor,
        },
    )


@require_POST
@login_required
def provisioning_catalog_request_bulk(request):
    """
    Create multiple ProvisioningRequest rows from selected service IDs.
    Uses acting user (manager can request on behalf of employee).
    """
    acting_user = _get_acting_user(request)

    # NEW: reason от модала
    reason = (request.POST.get("reason") or "").strip()

    service_ids = request.POST.getlist("service_ids")
    if not service_ids:
        messages.error(request, "No services selected.")
        return redirect("portal:provisioning_catalog")

    try:
        service_ids_int = [int(x) for x in service_ids]
    except ValueError:
        messages.error(request, "Invalid selection.")
        return redirect("portal:provisioning_catalog")

    assigned_ids = set(
        ServiceAssignment.objects.filter(user=acting_user, service_id__in=service_ids_int)
        .values_list("service_id", flat=True)
    )

    pending_ids = set(
        ProvisioningRequest.objects.filter(
            requester=acting_user,
            status=ProvisioningRequest.STATUS_PENDING,
            service_id__in=service_ids_int,
        ).values_list("service_id", flat=True)
    )

    to_create_ids = [sid for sid in service_ids_int if sid not in assigned_ids and sid not in pending_ids]
    if not to_create_ids:
        messages.info(request, "Nothing to request (already assigned or pending).")
        return redirect("portal:provisioning_my_requests")

    created = 0
    skipped_inactive = 0
    skipped_vendor_closed = 0

    with transaction.atomic():
        services = Service.objects.filter(id__in=to_create_ids).select_related("vendor")

        for svc in services:
            if hasattr(Service, "is_active") and not getattr(svc, "is_active", True):
                skipped_inactive += 1
                continue
            if hasattr(Vendor, "is_active") and svc.vendor and not getattr(svc.vendor, "is_active", True):
                skipped_vendor_closed += 1
                continue

            try:
                ProvisioningRequest.objects.create(
                    requester=acting_user,
                    service=svc,
                    status=ProvisioningRequest.STATUS_PENDING,
                    reason=reason,
                )
                created += 1
            except IntegrityError:
                # unique pending constraint
                pass

    if created:
        if acting_user.pk != request.user.pk:
            messages.success(request, f"Submitted {created} request(s) for {acting_user.username}.")
        else:
            messages.success(request, f"Submitted {created} request(s).")
    if skipped_inactive or skipped_vendor_closed:
        messages.info(request, "Some services were skipped because they are not available (inactive/closed vendor).")

    return redirect("portal:provisioning_my_requests")


@login_required
def provisioning_my_requests(request):
    """
    My Requests page (real data).
    Uses acting user.
    Supports status filter via ?status=pending|approved|rejected|cancelled
    """
    acting_user = _get_acting_user(request)

    # какво е избрано от табчетата горе
    status_key = (request.GET.get("status") or "").strip().lower()

    # map от ключ за URL към реалните стойности в модела
    status_filters = {
        "pending":   ProvisioningRequest.STATUS_PENDING,
        "approved":  ProvisioningRequest.STATUS_APPROVED,
        "rejected":  ProvisioningRequest.STATUS_REJECTED,
    }
    if hasattr(ProvisioningRequest, "STATUS_CANCELLED"):
        status_filters["cancelled"] = ProvisioningRequest.STATUS_CANCELLED

    reqs = (
        ProvisioningRequest.objects
        .filter(requester=acting_user)
        .select_related("service", "service__vendor", "decided_by")
        .order_by("-created_at", "-id")
    )

    # ако има валиден филтър – прилагаме го
    db_status = status_filters.get(status_key)
    if db_status:
        reqs = reqs.filter(status=db_status)

    return render(
        request,
        "portal/provisioning_my_requests.html",
        {
            "acting_user": acting_user,
            "is_acting": (acting_user.pk != request.user.pk),
            "requests": reqs,
            "current_status": status_key,
            # за шаблона – да не пишем 'approved' на ръка
            "status_pending": ProvisioningRequest.STATUS_PENDING,
            "status_approved": ProvisioningRequest.STATUS_APPROVED,
            "status_rejected": ProvisioningRequest.STATUS_REJECTED,
            "status_cancelled": getattr(
                ProvisioningRequest, "STATUS_CANCELLED", None
            ),
        },
    )


@login_required
def provisioning_approvals(request):
    """
    Provisioning Hub - Approvals (admin queue).
    Shows pending requests only.
    """
    is_prov_admin = _is_prov_admin(request.user)
    if not is_prov_admin:
        messages.error(request, "You do not have permission to access approvals.")
        return redirect("portal:provisioning_hub")

    approvals = (
        ProvisioningRequest.objects
        .select_related("requester", "service", "service__vendor")
        .filter(status=ProvisioningRequest.STATUS_PENDING)
        .order_by("-created_at", "-id")
    )

    return render(
        request,
        "portal/provisioning_approvals.html",
        {
            "is_prov_admin": is_prov_admin,
            "approvals": approvals,
        },
    )


@login_required
def provisioning_request_create(request, service_pk: int):
    """
    Single-service request create (kept for compatibility).
    Uses acting user.
    """
    acting_user = _get_acting_user(request)

    service = get_object_or_404(
        Service.objects.select_related("vendor"),
        pk=service_pk,
    )

    # Guardrails: service / vendor must be active
    if hasattr(Service, "is_active") and not getattr(service, "is_active", True):
        messages.error(request, "This service is closed and cannot be requested.")
        return redirect("portal:provisioning_catalog")

    if hasattr(Vendor, "is_active") and service.vendor and not getattr(service.vendor, "is_active", True):
        messages.error(request, "This vendor is closed and cannot be requested.")
        return redirect("portal:provisioning_catalog")

    # Already has access?
    if ServiceAssignment.objects.filter(user=acting_user, service=service).exists():
        messages.info(request, "Access already exists for this service.")
        return redirect("portal:provisioning_hub")

    if request.method == "POST":
        # 1) вече има pending?
        existing_pending = ProvisioningRequest.objects.filter(
            requester=acting_user,
            service=service,
            status=ProvisioningRequest.STATUS_PENDING,
        ).first()
        if existing_pending:
            messages.info(request, "A pending request already exists for this service.")
            return redirect("portal:provisioning_my_requests")

        # 2) reason – задължително
        reason = (request.POST.get("reason") or "").strip()
        if not reason:
            messages.error(request, "Please provide a short business reason for this request.")
            return render(
        request,
        "portal/provisioning_request_create.html",
        {
            "acting_user": acting_user,
            "is_acting": (acting_user.pk != request.user.pk),
            "service": service,
            "reason": reason,
        },
    )

        # 3) създаваме request с reason
        try:
            ProvisioningRequest.objects.create(
                requester=acting_user,
                service=service,
                status=ProvisioningRequest.STATUS_PENDING,
                reason=reason,
            )
            if acting_user.pk != request.user.pk:
                messages.success(
                    request,
                    f"Request submitted for {acting_user.username}: "
                    f"{service.vendor.name} – {service.name}.",
                )
            else:
                messages.success(
                    request,
                    f"Request submitted: {service.vendor.name} – {service.name}.",
                )
        except IntegrityError:
            messages.info(request, "A pending request already exists for this service.")

        return redirect("portal:provisioning_my_requests")

    # GET – първоначално зареждане
    return render(
        request,
        "portal/provisioning_request_create.html",
        {
            "acting_user": acting_user,
            "is_acting": (acting_user.pk != request.user.pk),
            "service": service,
            "reason": "",
        },
    )


@login_required
def provisioning_approval_decide(request, pk: int):
    """
    Keep single-item endpoint for backward compatibility.
    """
    if not _is_prov_admin(request.user):
        messages.error(request, "You do not have permission to access approvals.")
        return redirect("portal:provisioning_hub")

    if request.method != "POST":
        messages.error(request, "POST required.")
        return redirect("portal:provisioning_approvals")

    decision = (request.POST.get("decision") or "").strip().lower()
    if decision not in ("approve", "reject"):
        messages.error(request, "Invalid decision.")
        return redirect("portal:provisioning_approvals")

    pr = get_object_or_404(ProvisioningRequest, pk=pk)

    if pr.status != ProvisioningRequest.STATUS_PENDING:
        messages.warning(request, "This request is no longer pending.")
        return redirect("portal:provisioning_approvals")

    with transaction.atomic():
        pr = ProvisioningRequest.objects.select_for_update().get(pk=pk)

        if pr.status != ProvisioningRequest.STATUS_PENDING:
            messages.warning(request, "This request is no longer pending.")
            return redirect("portal:provisioning_approvals")

        pr.decided_at = timezone.now()
        pr.decided_by = request.user

        if decision == "approve":
            pr.status = ProvisioningRequest.STATUS_APPROVED
            pr.save(update_fields=["status", "decided_at", "decided_by"])
            ServiceAssignment.objects.get_or_create(
                user=pr.requester,
                service=pr.service,
                defaults={"assigned_by": request.user},
            )
        else:
            pr.status = ProvisioningRequest.STATUS_REJECTED
            pr.save(update_fields=["status", "decided_at", "decided_by"])

    messages.success(request, f"Decision recorded: {decision}.")
    return redirect("portal:provisioning_approvals")


@login_required
def provisioning_approvals_decide_bulk(request):
    """
    Bulk approve/reject selected pending requests.
    POST:
      - ids: list of ProvisioningRequest ids
      - decision: approve|reject
      - decision_note: optional text, записва се върху всички
    """
    if request.method != "POST":
        messages.error(request, "POST required.")
        return redirect("portal:provisioning_approvals")

    if not _is_prov_admin(request.user):
        messages.error(request, "You do not have permission to access approvals.")
        return redirect("portal:provisioning_hub")

    decision = (request.POST.get("decision") or "").strip().lower()
    if decision not in ("approve", "reject"):
        messages.error(request, "Invalid decision.")
        return redirect("portal:provisioning_approvals")

    # избрани заявки
    raw_ids = request.POST.getlist("ids")
    ids = [int(i) for i in raw_ids if str(i).isdigit()]
    if not ids:
        messages.warning(request, "No requests selected.")
        return redirect("portal:provisioning_approvals")

    decision_note = (request.POST.get("decision_note") or "").strip()

    processed = 0

    with transaction.atomic():
        qs = (
            ProvisioningRequest.objects
            .select_related("service", "service__vendor", "requester")
            .select_for_update()
            .filter(id__in=ids, status=ProvisioningRequest.STATUS_PENDING)
        )

        for pr in qs:
            pr.decided_at = timezone.now()
            pr.decided_by = request.user

            # ако имаме decision_note – запиши го
            if decision_note and hasattr(pr, "decision_note"):
                pr.decision_note = decision_note

            if decision == "approve":
                pr.status = ProvisioningRequest.STATUS_APPROVED
                pr.save(
                    update_fields=[
                        "status",
                        "decided_at",
                        "decided_by",
                        *(["decision_note"] if decision_note and hasattr(pr, "decision_note") else []),
                    ]
                )

                # създай assignment, ако още го няма
                ServiceAssignment.objects.get_or_create(
                    user=pr.requester,
                    service=pr.service,
                    defaults={"assigned_by": request.user},
                )
            else:
                pr.status = ProvisioningRequest.STATUS_REJECTED
                pr.save(
                    update_fields=[
                        "status",
                        "decided_at",
                        "decided_by",
                        *(["decision_note"] if decision_note and hasattr(pr, "decision_note") else []),
                    ]
                )

            processed += 1

    if processed:
        messages.success(request, f"{processed} request(s) {decision}d.")
    else:
        messages.warning(request, "No pending requests matched your selection.")

    return redirect("portal:provisioning_approvals")

    
@login_required
@require_POST
def provisioning_access_remove(request, service_pk: int):
    """
    Remove access to a service (deletes ServiceAssignment).
    Uses acting user, so a manager can remove access for direct report.
    """
    acting_user = _get_acting_user(request)

    assignment = ServiceAssignment.objects.filter(
        user=acting_user,
        service_id=service_pk,
    ).first()

    if not assignment:
        messages.error(request, "No active assignment found for this service.")
        return redirect("portal:provisioning_hub")

    assignment.delete()

    if acting_user.pk != request.user.pk:
        messages.success(request, f"Access removed for {acting_user.username}.")
    else:
        messages.success(request, "Access removed.")

    return redirect("portal:provisioning_hub")


@login_required
def report_center(request):
    """
    Reports Center.

    - view=overview               -> картите с наличните репорти
    - view=users_cost             -> Users · access cost
    - view=services_catalog       -> Services catalog (pricing)
    - view=contracts_renewals     -> Contracts renewals schedule
    - view=vendor_spend_year      -> Vendor spend by year (Invoice-based)
    - view=user_activity_timeline -> User activity timeline (logins + access)
    - view=builder                -> generic табличен report builder
    """
    active_view = (request.GET.get("view") or "overview").strip() or "overview"
    if active_view not in {
        "overview",
        "users_cost",
        "services_catalog",
        "contracts_renewals",
        "vendor_spend_year",
        "user_activity_timeline",
        "builder",
    }:
        active_view = "overview"

    # ---------------- общи променливи ----------------
    user_cost_rows: list[dict] = []
    services_catalog_rows: list[dict] = []
    contracts_renewals_rows: list[dict] = []
    vendor_spend_rows: list[dict] = []
    user_activity_rows: list[dict] = []

    # builder
    builder_datasets = [
        {"key": "users_profiles", "label": "Users & profiles"},
        {"key": "user_access", "label": "User · services"},
    ]
    builder_active_dataset = (
        request.GET.get("dataset") or "users_profiles"
    ).strip() or "users_profiles"
    valid_dataset_keys = {d["key"] for d in builder_datasets}
    if builder_active_dataset not in valid_dataset_keys:
        builder_active_dataset = "users_profiles"

    builder_columns: list[dict] = []
    builder_rows: list[dict] = []
    builder_selected_cols: set[str] = set()
    builder_filters: dict[str, str] = {}
    builder_preview_limit = 500
    builder_total_count = 0
    builder_preview_count = 0

    # ============================================================
    # 1) Users · access cost
    # ============================================================
    if active_view == "users_cost":
        assignments = (
            ServiceAssignment.objects
            .select_related(
                "user",
                "user__profile",
                "user__profile__cost_center",
                "service",
                "service__vendor",
            )
            .filter(user__is_active=True)
        )

        per_user: dict[int, dict] = {}

        for a in assignments:
            user = a.user
            service = a.service
            if not user or not service:
                continue

            entry = per_user.get(user.pk)
            if not entry:
                profile = getattr(user, "profile", None)
                cost_center = getattr(profile, "cost_center", None)

                # Full name: profile.full_name > Django full_name > username
                if profile and getattr(profile, "full_name", ""):
                    full_name = profile.full_name
                else:
                    fn = getattr(user, "get_full_name", lambda: "")()
                    full_name = fn or user.username

                entry = {
                    "user": user,
                    "username": user.username,
                    "full_name": full_name,
                    "cost_center": cost_center,
                    "services_count": 0,
                    "total_cost": Decimal("0"),
                    "currencies": set(),
                }
                per_user[user.pk] = entry

            entry["services_count"] += 1

            price = getattr(service, "list_price", None)
            currency = getattr(service, "default_currency", "") or ""
            if price is not None:
                try:
                    entry["total_cost"] += price
                except Exception:
                    pass
            if currency:
                entry["currencies"].add(currency)

        for entry in per_user.values():
            currencies = entry["currencies"]
            if not currencies:
                currency_label = ""
            elif len(currencies) == 1:
                currency_label = list(currencies)[0]
            else:
                currency_label = "Mixed"

            user_cost_rows.append({
                "user": entry["user"],
                "username": entry["username"],
                "full_name": entry["full_name"],
                "cost_center": entry["cost_center"],
                "services_count": entry["services_count"],
                "total_cost": entry["total_cost"],
                "currency": currency_label,
            })

        user_cost_rows.sort(key=lambda r: r["username"].lower())

        # CSV export
        if (request.GET.get("export") or "").lower() == "csv":
            headers = [
                "username",
                "full_name",
                "cost_center_code",
                "cost_center_name",
                "services_count",
                "total_cost",
                "currency",
            ]
            rows = []
            for r in user_cost_rows:
                cc = r["cost_center"]
                rows.append([
                    r["username"],
                    r["full_name"],
                    getattr(cc, "code", "") if cc else "",
                    getattr(cc, "name", "") if cc else "",
                    str(r["services_count"]),
                    str(r["total_cost"]) if r["total_cost"] is not None else "",
                    r["currency"],
                ])

            filename = (
                f"datanaut_report_users_cost_"
                f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            return _csv_response(filename, headers, rows)

    # ============================================================
    # 2) Services catalog (pricing)
    # ============================================================
    if active_view == "services_catalog":
        search = (request.GET.get("q") or "").strip()
        status = (request.GET.get("status") or "").strip().lower()

        qs = Service.objects.select_related("vendor")

        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(service_code__icontains=search)
                | Q(vendor__name__icontains=search)
                | Q(category__icontains=search)
            )

        if status == "active":
            qs = qs.filter(is_active=True)
        elif status == "closed":
            qs = qs.filter(is_active=False)

        qs = qs.order_by("vendor__name", "name")

        for s in qs:
            vendor = getattr(s, "vendor", None)
            services_catalog_rows.append({
                "service_name": getattr(s, "name", "") or "",
                "service_code": getattr(s, "service_code", "") or "",
                "vendor_name": getattr(vendor, "name", "") if vendor else "",
                "category": getattr(s, "category", "") or "",
                "status": "Active" if getattr(s, "is_active", True) else "Closed",
                "list_price": getattr(s, "list_price", None),
                "currency": getattr(s, "default_currency", "") or "",
                "billing_period": (
                    getattr(s, "billing_period", "")
                    or getattr(s, "default_billing_frequency", "")
                    or getattr(s, "billing_frequency", "")
                ),
            })

        # CSV export
        if (request.GET.get("export") or "").lower() == "csv":
            headers = [
                "service_name",
                "service_code",
                "vendor_name",
                "category",
                "status",
                "list_price",
                "currency",
                "billing_period",
            ]
            rows = []
            for r in services_catalog_rows:
                rows.append([
                    r["service_name"],
                    r["service_code"],
                    r["vendor_name"],
                    r["category"],
                    r["status"],
                    str(r["list_price"]) if r["list_price"] is not None else "",
                    r["currency"],
                    r["billing_period"],
                ])

            filename = (
                f"datanaut_report_services_catalog_"
                f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            return _csv_response(filename, headers, rows)

    # ============================================================
    # 3) Contracts renewals schedule
    # ============================================================
    if active_view == "contracts_renewals":
        today = date.today()

        qs = Contract.objects.select_related("vendor").all()

        for c in qs:
            vendor = getattr(c, "vendor", None)

            end_date = (
                getattr(c, "end_date", None)
                or getattr(c, "valid_to", None)
                or getattr(c, "renewal_date", None)
            )
            start_date = (
                getattr(c, "start_date", None)
                or getattr(c, "valid_from", None)
            )

            annual_value = (
                getattr(c, "annual_value", None)
                or getattr(c, "contract_value", None)
            )
            risk_flag = (
                getattr(c, "risk_flag", "")
                or getattr(c, "risk_level", "")
                or ""
            )

            days_to_renewal = None
            if isinstance(end_date, (datetime, date)):
                end_date_date = end_date.date() if isinstance(end_date, datetime) else end_date
                days_to_renewal = (end_date_date - today).days

            contracts_renewals_rows.append({
                "contract_code": (
                    getattr(c, "reference", "")
                    or getattr(c, "code", "")
                    or str(getattr(c, "id", ""))
                ),
                "service_name": getattr(getattr(c, "service", None), "name", ""),
                "vendor_name": getattr(vendor, "name", "") if vendor else "",
                "legal_entity": getattr(c, "legal_entity", "") or "",
                "start_date": start_date,
                "end_date": end_date,
                "status": getattr(c, "status", "") or "",
                "annual_value": annual_value,
                "currency": getattr(c, "currency", "") or "",
                "risk_flag": risk_flag,
                "days_to_renewal": days_to_renewal,
            })

        def _sort_key(r: dict):
            end = r["end_date"]
            if isinstance(end, datetime):
                end = end.date()
            if isinstance(end, date):
                return (0, end)
            return (1, date.max)

        contracts_renewals_rows.sort(key=_sort_key)

        # CSV export
        if (request.GET.get("export") or "").lower() == "csv":
            headers = [
                "contract_code",
                "service_name",
                "vendor_name",
                "legal_entity",
                "start_date",
                "end_date",
                "status",
                "annual_value",
                "currency",
                "risk_flag",
                "days_to_renewal",
            ]
            rows = []
            for r in contracts_renewals_rows:
                rows.append([
                    r["contract_code"],
                    r["service_name"],
                    r["vendor_name"],
                    r["legal_entity"],
                    r["start_date"].isoformat()
                    if isinstance(r["start_date"], (date, datetime)) else "",
                    r["end_date"].isoformat()
                    if isinstance(r["end_date"], (date, datetime)) else "",
                    r["status"],
                    str(r["annual_value"]) if r["annual_value"] is not None else "",
                    r["currency"],
                    r["risk_flag"],
                    "" if r["days_to_renewal"] is None else str(r["days_to_renewal"]),
                ])

            filename = (
                f"datanaut_report_contracts_renewals_"
                f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            return _csv_response(filename, headers, rows)

    # ============================================================
    # 4) Vendor spend by year (Invoice-based)
    # ============================================================
    if active_view == "vendor_spend_year":
        amount_field = _first_existing_field(
            Invoice,
            ["total_amount", "amount", "net_amount", "gross_amount", "value"],
        )
        date_field = _first_existing_field(
            Invoice,
            ["invoice_date", "date", "issue_date", "period_start", "period_end"],
        )
        currency_field = _first_existing_field(
            Invoice,
            ["currency", "currency_code"],
        )

        if amount_field and date_field and currency_field:
            qs = (
                Invoice.objects
                .filter(owner=request.user)
                .select_related("vendor")
                .annotate(year=ExtractYear(date_field))
                .values("year", "vendor__name", currency_field)
                .annotate(total_spend=Sum(amount_field))
                .order_by("-year", "vendor__name")
            )

            for row in qs:
                vendor_spend_rows.append({
                    "year": row.get("year"),
                    "vendor_name": row.get("vendor__name") or "",
                    "currency": row.get(currency_field) or "",
                    "total_spend": row.get("total_spend") or Decimal("0"),
                })

        # CSV export
        if (request.GET.get("export") or "").lower() == "csv":
            headers = ["year", "vendor_name", "currency", "total_spend"]
            rows = []
            for r in vendor_spend_rows:
                rows.append([
                    str(r["year"]) if r["year"] is not None else "",
                    r["vendor_name"],
                    r["currency"],
                    str(r["total_spend"]) if r["total_spend"] is not None else "",
                ])
            filename = (
                f"datanaut_report_vendor_spend_year_"
                f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            return _csv_response(filename, headers, rows)

    # ============================================================
    # 5) User activity timeline
    # ============================================================
    if active_view == "user_activity_timeline":
        today = date.today()
        recent_window_days = 90
        dormant_days = 60
        recent_threshold = today - timedelta(days=recent_window_days)
        dormant_threshold = today - timedelta(days=dormant_days)

        users_qs = (
            User.objects
            .select_related("profile")
            .order_by("username")
        )

        for u in users_qs:
            profile = getattr(u, "profile", None)

            if profile and getattr(profile, "full_name", ""):
                full_name = profile.full_name
            else:
                fn = getattr(u, "get_full_name", lambda: "")()
                full_name = fn or u.username

            assignments = (
                ServiceAssignment.objects
                .select_related("service", "service__vendor")
                .filter(user=u)
            )
            service_names = []
            for a in assignments:
                s = getattr(a, "service", None)
                if s and getattr(s, "name", ""):
                    service_names.append(s.name)
            services_summary = ", ".join(sorted(set(service_names)))

            last_login = u.last_login
            last_activity_date = last_login.date() if last_login else None

            if last_activity_date and last_activity_date >= recent_threshold:
                active_days_90d = (today - last_activity_date).days
            elif last_activity_date:
                active_days_90d = 0
            else:
                active_days_90d = None

            if last_activity_date and last_activity_date <= dormant_threshold:
                dormant_since = last_activity_date
            else:
                dormant_since = None

            user_activity_rows.append({
                "username": u.username,
                "full_name": full_name,
                "last_activity": last_login,
                "active_days_90d": active_days_90d,
                "dormant_since": dormant_since,
                "services_summary": services_summary,
            })

        def _ua_sort_key(row):
            la = row["last_activity"]
            if la is None:
                return (2, "")  # никога не е логвал
            return (0, la) if row["dormant_since"] else (1, la)

        user_activity_rows.sort(key=_ua_sort_key, reverse=True)

        # CSV export
        if (request.GET.get("export") or "").lower() == "csv":
            headers = [
                "username",
                "full_name",
                "last_activity",
                "active_days_90d",
                "dormant_since",
                "services_summary",
            ]
            rows = []
            for r in user_activity_rows:
                rows.append([
                    r["username"],
                    r["full_name"],
                    r["last_activity"].isoformat() if r["last_activity"] else "",
                    "" if r["active_days_90d"] is None else str(r["active_days_90d"]),
                    r["dormant_since"].isoformat() if r["dormant_since"] else "",
                    r["services_summary"],
                ])
            filename = (
                f"datanaut_report_user_activity_timeline_"
                f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            return _csv_response(filename, headers, rows)

    # ============================================================
    # 6) Report builder (generic datasets)
    # ============================================================
    if active_view == "builder":
        # 6.1 Users & profiles
        if builder_active_dataset == "users_profiles":
            builder_columns = [
                {"key": "username",         "label": "Username"},
                {"key": "full_name",        "label": "Full name"},
                {"key": "email",            "label": "Email"},
                {"key": "status",           "label": "Status"},
                {"key": "cost_center_code", "label": "Cost center code"},
                {"key": "cost_center_name", "label": "Cost center name"},
                {"key": "manager",          "label": "Manager"},
                {"key": "location",         "label": "Location"},
                {"key": "legal_entity",     "label": "Legal entity"},
            ]

            qs = (
                User.objects
                .select_related("profile", "profile__cost_center", "profile__manager")
                .order_by("username")
            )

            for u in qs:
                profile = getattr(u, "profile", None)
                cc = getattr(profile, "cost_center", None)

                if profile and getattr(profile, "full_name", ""):
                    full_name = profile.full_name
                else:
                    fn = getattr(u, "get_full_name", lambda: "")()
                    full_name = fn or ""

                builder_rows.append({
                    "username": u.username,
                    "full_name": full_name,
                    "email": u.email or "",
                    "status": "Active" if u.is_active else "Closed",
                    "cost_center_code": getattr(cc, "code", "") if cc else "",
                    "cost_center_name": getattr(cc, "name", "") if cc else "",
                    "manager": getattr(
                        getattr(profile, "manager", None),
                        "username",
                        "",
                    ) if profile else "",
                    "location": getattr(profile, "location", "") if profile else "",
                    "legal_entity": getattr(profile, "legal_entity", "") if profile else "",
                })

        # 6.2 User · services
        elif builder_active_dataset == "user_access":
            builder_columns = [
                {"key": "username",         "label": "Username"},
                {"key": "full_name",        "label": "Full name"},
                {"key": "email",            "label": "Email"},
                {"key": "user_status",      "label": "User status"},
                {"key": "service_name",     "label": "Service"},
                {"key": "vendor_name",      "label": "Vendor"},
                {"key": "service_category", "label": "Service category"},
                {"key": "service_status",   "label": "Service status"},
                {"key": "list_price",       "label": "List price"},
                {"key": "currency",         "label": "Currency"},
                {"key": "cost_center_code", "label": "Cost center code"},
                {"key": "cost_center_name", "label": "Cost center name"},
            ]

            qs = (
                ServiceAssignment.objects
                .select_related(
                    "user",
                    "user__profile",
                    "user__profile__cost_center",
                    "service",
                    "service__vendor",
                )
                .order_by("user__username", "service__vendor__name", "service__name")
            )

            for a in qs:
                u = a.user
                s = a.service
                if not u or not s:
                    continue

                profile = getattr(u, "profile", None)
                cc = getattr(profile, "cost_center", None)
                vendor = getattr(s, "vendor", None)

                if profile and getattr(profile, "full_name", ""):
                    full_name = profile.full_name
                else:
                    fn = getattr(u, "get_full_name", lambda: "")()
                    full_name = fn or ""

                builder_rows.append({
                    "username": u.username,
                    "full_name": full_name,
                    "email": u.email or "",
                    "user_status": "Active" if u.is_active else "Closed",
                    "service_name": s.name or "",
                    "vendor_name": getattr(vendor, "name", "") if vendor else "",
                    "service_category": getattr(s, "category", "") or "",
                    "service_status": "Active" if getattr(s, "is_active", True) else "Closed",
                    "list_price": str(getattr(s, "list_price", "") or ""),
                    "currency": getattr(s, "default_currency", "") or "",
                    "cost_center_code": getattr(cc, "code", "") if cc else "",
                    "cost_center_name": getattr(cc, "name", "") if cc else "",
                })

        # избрани колони
        builder_selected_cols = set(request.GET.getlist("col"))
        if not builder_selected_cols and builder_columns:
            builder_selected_cols = {c["key"] for c in builder_columns}

        # per-column филтри f_<colkey>
        for col in builder_columns:
            raw = (request.GET.get(f"f_{col['key']}") or "").strip()
            if raw:
                builder_filters[col["key"]] = raw

        def _row_matches(row: dict) -> bool:
            for k, search in builder_filters.items():
                val = str(row.get(k, "") or "")
                if search.lower() not in val.lower():
                    return False
            return True

        filtered_rows = [r for r in builder_rows if _row_matches(r)]
        builder_total_count = len(filtered_rows)

        # CSV export
        if (request.GET.get("export") or "").lower() == "csv":
            headers = [
                c["label"]
                for c in builder_columns
                if c["key"] in builder_selected_cols
            ]
            rows = []
            for r in filtered_rows:
                rows.append([
                    str(r.get(c["key"], "") or "")
                    for c in builder_columns
                    if c["key"] in builder_selected_cols
                ])

            filename = (
                f"datanaut_report_builder_{builder_active_dataset}_"
                f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            return _csv_response(filename, headers, rows)

        preview_rows = filtered_rows[:builder_preview_limit]
        builder_preview_count = len(preview_rows)
        builder_rows = preview_rows

    # ============================================================
    # 7) Картите за OVERVIEW
    # ============================================================
    base_url = reverse("portal:reports")
    available_reports = [
        {
            "code": "users_cost",
            "name": "Users · access cost",
            "description": (
                "Per-user view of assigned services and estimated run-rate "
                "based on list prices."
            ),
            "url": f"{base_url}?view=users_cost",
        },
        {
            "code": "services_catalog",
            "name": "Services catalog (pricing)",
            "description": "Flat catalog of services with vendor and list price.",
            "url": f"{base_url}?view=services_catalog",
        },
        {
            "code": "contracts_renewals",
            "name": "Contracts renewals schedule",
            "description": (
                "Upcoming contract renewals with annual value and risk flags."
            ),
            "url": f"{base_url}?view=contracts_renewals",
        },
        {
            "code": "vendor_spend_year",
            "name": "Vendor spend by year",
            "description": (
                "Yearly spend per vendor based on invoice history; can power "
                "dashboard spend widgets."
            ),
            "url": f"{base_url}?view=vendor_spend_year",
        },
        {
            "code": "user_activity_timeline",
            "name": "User activity timeline",
            "description": (
                "Login and usage signals per user; supports dormant/active "
                "usage views."
            ),
            "url": f"{base_url}?view=user_activity_timeline",
        },
        {
            "code": "builder",
            "name": "Report builder",
            "description": (
                "Self-service tabular view over users, services and access. "
                "Configure columns and export CSV."
            ),
            "url": f"{base_url}?view=builder",
        },
    ]

    return render(
        request,
        "portal/reports.html",
        {
            "available_reports": available_reports,
            "active_view": active_view,
            "user_cost_rows": user_cost_rows,
            "services_catalog_rows": services_catalog_rows,
            "contracts_renewals_rows": contracts_renewals_rows,
            "vendor_spend_rows": vendor_spend_rows,
            "user_activity_rows": user_activity_rows,
            # builder context
            "builder_datasets": builder_datasets,
            "builder_active_dataset": builder_active_dataset,
            "builder_columns": builder_columns,
            "builder_rows": builder_rows,
            "builder_selected_cols": builder_selected_cols,
            "builder_filters": builder_filters,
            "builder_preview_limit": builder_preview_limit,
            "builder_total_count": builder_total_count,
            "builder_preview_count": builder_preview_count,
        },
    )


# ----------
# USAGE
# ----------

def _build_usage_snapshot():
    """
    Общ usage snapshot, който ползваме за:
      - overview (desks)
      - vendor inventory
      - user inventory
    """
    UserModel = get_user_model()
    for u in UserModel.objects.all().iterator():
        UserProfile.objects.get_or_create(user=u)

    now = timezone.now()
    dormant_threshold_days = 60
    window_90d = now - timedelta(days=90)

    assignments = (
        ServiceAssignment.objects
        .select_related(
            "user",
            "user__profile",
            "user__profile__cost_center",
            "service",
            "service__vendor",
        )
    )

    desks_data: dict[int, dict] = {}
    all_vendors = set()
    dormant_users_map: dict[int, dict] = {}
    vendor_stats: dict[int, dict] = {}
    user_stats: dict[int, dict] = {}

    total_licences = 0
    total_dormant_licences = 0
    potential_savings = Decimal("0")

    for a in assignments:
        user = a.user
        svc = a.service
        vendor = getattr(svc, "vendor", None)
        profile = getattr(user, "profile", None)
        cc = getattr(profile, "cost_center", None)

        if not cc:
            continue

        desk_key = cc.pk
        if cc.code and cc.name:
            desk_label = f"{cc.code} – {cc.name}"
        else:
            desk_label = cc.code or cc.name or "Unmapped"

        vendor_name = vendor.name if vendor else "—"
        all_vendors.add(vendor_name)

        list_price = getattr(svc, "list_price", None) or Decimal("0")
        category = getattr(svc, "category", "") or ""

        last_login = user.last_login
        days_since_login = None
        is_dormant = False
        is_recent = False

        if last_login:
            days_since_login = (now.date() - last_login.date()).days
            is_dormant = days_since_login > dormant_threshold_days
            is_recent = last_login >= window_90d
        else:
            is_dormant = True

        # ----- per desk -----
        d = desks_data.setdefault(
            desk_key,
            {
                "desk_label": desk_label,
                "vendor_names": set(),
                "licences": 0,
                "dormant_licences": 0,
                "recent_users": set(),
                "all_users": set(),
                "total_price": Decimal("0"),
                "dormant_price": Decimal("0"),
                "vendor_by_category": defaultdict(set),
            },
        )

        d["vendor_names"].add(vendor_name)
        d["licences"] += 1
        d["all_users"].add(user.pk)
        d["total_price"] += list_price
        d["vendor_by_category"][category].add(vendor_name)

        if is_recent:
            d["recent_users"].add(user.pk)

        if is_dormant:
            d["dormant_licences"] += 1
            d["dormant_price"] += list_price

            du = dormant_users_map.get(user.pk)
            prev_days = du["days_since_login"] if du else None
            curr_days = days_since_login if days_since_login is not None else 9999
            if prev_days is None or curr_days > prev_days:
                dormant_users_map[user.pk] = {
                    "username": user.username,
                    "desk_label": desk_label,
                    "vendor": vendor_name,
                    "days_since_login": days_since_login,
                }

        total_licences += 1
        if is_dormant:
            total_dormant_licences += 1
            potential_savings += list_price

        # ----- per vendor -----
        if vendor:
            v = vendor_stats.setdefault(
                vendor.pk,
                {
                    "vendor": vendor,
                    "vendor_name": vendor_name,
                    "licences": 0,
                    "dormant_licences": 0,
                    "total_price": Decimal("0"),
                    "dormant_price": Decimal("0"),
                    "desks": set(),
                },
            )
            v["licences"] += 1
            v["total_price"] += list_price
            v["desks"].add(desk_label)
            if is_dormant:
                v["dormant_licences"] += 1
                v["dormant_price"] += list_price

        # ----- per user -----
        ustat = user_stats.setdefault(
            user.pk,
            {
                "user": user,
                "username": user.username,
                "desk_label": desk_label,
                "services": 0,
                "dormant_services": 0,
                "total_price": Decimal("0"),
                "last_login": None,
            },
        )
        ustat["services"] += 1
        ustat["total_price"] += list_price
        if is_dormant:
            ustat["dormant_services"] += 1
        if last_login and (ustat["last_login"] is None or last_login > ustat["last_login"]):
            ustat["last_login"] = last_login

    # ---- KPIs ----
    if total_licences > 0:
        healthy_licences = total_licences - total_dormant_licences
        healthy_percent = int(round(100 * healthy_licences / total_licences))
    else:
        healthy_percent = 0

    vendors_count = len(all_vendors)
    desks_count = len(desks_data)

    # ---- desks table ----
    desk_rows = []
    overlapping_desks = 0
    severity_order = {"high": 0, "medium": 1, "low": 2}

    for d in desks_data.values():
        licences = d["licences"]
        dormant = d["dormant_licences"]
        recent_users = len(d["recent_users"])
        all_users = len(d["all_users"])
        vendors = d["vendor_names"]

        dormant_ratio = (dormant / licences) if licences else 0
        recent_ratio = (recent_users / all_users) if all_users else 0

        has_overlap = any(len(vs) > 1 for vs in d["vendor_by_category"].values())
        if has_overlap:
            overlapping_desks += 1

        if len(vendors) == 1:
            vendor_label = next(iter(vendors))
        elif len(vendors) == 0:
            vendor_label = "—"
        else:
            vendor_label = "Mixed vendors"

        if dormant >= 3 and dormant_ratio >= 0.3:
            severity = "high"
            usage_signal = f"{dormant} dormant licences"
            last_90_days_text = "No/low logins in last 90 days"
        elif has_overlap:
            severity = "medium"
            usage_signal = "Overlap across vendors"
            last_90_days_text = "Multiple vendors with similar coverage"
        elif dormant > 0:
            severity = "medium"
            usage_signal = f"{dormant} low-usage licences"
            last_90_days_text = "Mixed usage across users"
        else:
            severity = "low"
            usage_signal = "Healthy usage"
            last_90_days_text = (
                "Stable, >90% active users" if recent_ratio >= 0.9 else "Mostly active users"
            )

        desk_rows.append(
            {
                "desk_label": d["desk_label"],
                "vendor_label": vendor_label,
                "licences": licences,
                "usage_signal": usage_signal,
                "last_90_days": last_90_days_text,
                "severity": severity,
                "severity_order": severity_order.get(severity, 3),
            }
        )

    desk_rows.sort(key=lambda r: (r["severity_order"], -r["licences"]))

    # ---- dormant users (side card) ----
    dormant_users = list(dormant_users_map.values())
    dormant_users.sort(key=lambda r: (r["days_since_login"] or 9999), reverse=True)

    # ---- overlapping products (side card) ----
    overlapping_rows = []
    for d in desks_data.values():
        desk_label = d["desk_label"]
        for category, vendors in d["vendor_by_category"].items():
            if len(vendors) < 2:
                continue
            vendors_label = " / ".join(sorted(vendors))
            if category:
                opportunity = f"Multiple vendors for {category} – review bundles."
            else:
                opportunity = "Multiple vendors for similar coverage – consider consolidation."
            overlapping_rows.append(
                {
                    "desk_label": desk_label,
                    "vendors_label": vendors_label,
                    "opportunity": opportunity,
                }
            )

    # ---- vendor inventory data ----
    vendor_rows = []
    for v in vendor_stats.values():
        licences = v["licences"]
        dormant = v["dormant_licences"]
        active = licences - dormant
        vendor_rows.append(
            {
                "vendor": v["vendor"],
                "vendor_name": v["vendor_name"],
                "licences": licences,
                "active_licences": active,
                "dormant_licences": dormant,
                "total_price": v["total_price"],
                "dormant_price": v["dormant_price"],
                "desks_count": len(v["desks"]),
            }
        )
    vendor_rows.sort(key=lambda r: r["vendor_name"].lower())

    # ---- user inventory data ----
    user_rows = []
    for u in user_stats.values():
        last_login = u["last_login"]
        if last_login:
            days_since = (now.date() - last_login.date()).days
        else:
            days_since = None
        is_dormant = u["dormant_services"] > 0 and (
            days_since is None or days_since > dormant_threshold_days
        )
        user_rows.append(
            {
                "user": u["user"],
                "username": u["username"],
                "desk_label": u["desk_label"],
                "services": u["services"],
                "dormant_services": u["dormant_services"],
                "total_price": u["total_price"],
                "last_login": last_login,
                "days_since_login": days_since,
                "status": "Dormant" if is_dormant else "Active",
            }
        )

    user_rows.sort(
        key=lambda r: (
            0 if r["status"] == "Dormant" else 1,
            -(r["services"]),
        )
    )

    kpis = {
        "licences_monitored": total_licences,
        "potential_savings": potential_savings,
        "healthy_percent": healthy_percent,
        "overlapping_desks": overlapping_desks,
        "vendors_count": vendors_count,
        "desks_count": desks_count,
    }

    return {
        "kpis": kpis,
        "desk_rows": desk_rows,
        "dormant_users": dormant_users,
        "overlapping_rows": overlapping_rows,
        "vendor_rows": vendor_rows,
        "user_rows": user_rows,
    }


@login_required
def usage_overview(request):
    """
    Licence usage overview, базиран на реалните:
    - ServiceAssignment (user × service)
    - User.last_login
    - UserProfile.cost_center (desk)
    - Service.vendor, Service.category, Service.list_price
    """
    snapshot = _build_usage_snapshot()
    desk_rows = snapshot["desk_rows"]

    # ---------- CSV export на desk таблицата ----------
    if (request.GET.get("export") or "").lower() == "csv":
        headers = [
            "desk_label",
            "vendor_label",
            "licences",
            "usage_signal",
            "last_90_days",
            "severity",
        ]
        rows = []
        for r in desk_rows:
            rows.append([
                r["desk_label"],
                r["vendor_label"],
                str(r["licences"]),
                r["usage_signal"],
                r["last_90_days"],
                r["severity"],
            ])

        filename = (
            f"datanaut_usage_desks_"
            f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        return _csv_response(filename, headers, rows)
    # ---------------------------------------------------

    context = {
        "kpis": snapshot["kpis"],
        "desk_rows": desk_rows,
        "dormant_users": snapshot["dormant_users"][:10],
        "overlapping_rows": snapshot["overlapping_rows"][:10],
    }
    return render(request, "portal/usage.html", context)

@login_required
def usage_contract(request):
    contracts = (
        Contract.objects.filter(owner=request.user)
        .select_related("vendor", "owning_cost_center")
        .order_by("-start_date", "-created_at")
    )

    # сумираме annual_value вместо несъществуващото total_value
    agg = contracts.aggregate(total_annual=Sum("annual_value"))
    total_annual = agg["total_annual"] or Decimal("0")

    contract_count = contracts.count()

    context = {
        "contracts": contracts,
        "contract_count": contract_count,
        "total_annual": total_annual,
        "active_tab": "contracts",   # за да свети правилния таб в менюто
    }
    return render(request, "portal/usage_contract.html", context)


@login_required
def usage_vendors(request):
    """
    Vendor inventory, базирано на общия usage snapshot.
    """
    snapshot = _build_usage_snapshot()
    vendor_rows = snapshot["vendor_rows"]
    kpis = snapshot["kpis"]

    show_closed = (request.GET.get("show_closed") in ("1", "true", "True", "on", "yes"))

    if not show_closed:
        filtered = []
        for row in vendor_rows:
            vendor = row.get("vendor")
            if hasattr(vendor, "is_active"):
                if vendor.is_active:
                    filtered.append(row)
            else:
                filtered.append(row)
        vendor_rows = filtered

    context = {
        "vendor_rows": vendor_rows,
        "kpis": kpis,
        "show_closed": show_closed,
        "active_tab": "vendors",
    }
    return render(request, "portal/usage_vendors.html", context)


@login_required
def usage_users(request):
    """
    User inventory – списък с потребители от общия usage snapshot,
    със сървърни филтри и контрол върху броя редове.
    """
    snapshot = _build_usage_snapshot()
    user_rows = snapshot["user_rows"]

    total_users = len(user_rows)

    # --- филтри от query string ---
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "all").lower()

    filtered = user_rows

    # статус филтър
    if status == "active":
        filtered = [u for u in filtered if u["status"] == "Active"]
    elif status == "dormant":
        filtered = [u for u in filtered if u["status"] == "Dormant"]

    # текстово търсене по user / desk
    if q:
        q_lower = q.lower()
        filtered = [
            u
            for u in filtered
            if q_lower in (u["username"] or "").lower()
            or q_lower in (u["desk_label"] or "").lower()
        ]

    filtered_count = len(filtered)

    # --- колко реда да показваме ---
    page_sizes = [10, 20, 30, 40, 50, 100]
    page_size_param = request.GET.get("limit") or request.GET.get("page_size") or "20"
    try:
        page_size = int(page_size_param)
        if page_size not in page_sizes:
            page_size = 20
    except (TypeError, ValueError):
        page_size = 20

    rows = filtered[:page_size]

    # малко KPI за side card
    dormant_users_count = len([u for u in user_rows if u["status"] == "Dormant"])
    dormant_percent = (
        int(round(100 * dormant_users_count / total_users)) if total_users else 0
    )

    context = {
        "user_rows": rows,
        "total_users": total_users,
        "filtered_count": filtered_count,
        "dormant_users_count": dormant_users_count,
        "dormant_percent": dormant_percent,
        "page_size": page_size,
        "page_sizes": page_sizes,
        "q": q,
        "status": status,
    }
    return render(request, "portal/usage_users.html", context)

@login_required
def usage_invoices(request):
    invoices = (
        Invoice.objects.filter(owner=request.user)
        .select_related("vendor", "contract")
        .order_by("-invoice_date", "-id")
    )

    # агрегации по реалните полета total_amount и tax_amount
    agg = invoices.aggregate(
        total_amount_sum=Sum("total_amount"),
        tax_amount_sum=Sum("tax_amount"),
    )

    total_amount = agg["total_amount_sum"] or Decimal("0")
    tax_amount = agg["tax_amount_sum"] or Decimal("0")
    invoice_count = invoices.count()

    context = {
        "invoices": invoices,
        "invoice_count": invoice_count,
        "total_amount": total_amount,
        "tax_amount": tax_amount,
        "active_tab": "invoices",  # за да светне правилния таб
    }
    return render(request, "portal/usage_invoices.html", context)


@login_required
def usage_vendors(request):
    """
    Vendor inventory, базирано на общия usage snapshot.
    """
    snapshot = _build_usage_snapshot()
    vendor_rows = snapshot["vendor_rows"]
    kpis = snapshot["kpis"]

    show_closed = (request.GET.get("show_closed") in ("1", "true", "True", "on", "yes"))

    # ако не искаме "затворени" доставчици, филтрираме по vendor.is_active
    if not show_closed:
        filtered = []
        for row in vendor_rows:
            vendor = row.get("vendor")
            # ако моделът няма is_active, приемаме че е активен
            if hasattr(vendor, "is_active"):
                if vendor.is_active:
                    filtered.append(row)
            else:
                filtered.append(row)
        vendor_rows = filtered

    context = {
        "vendor_rows": vendor_rows,
        "kpis": kpis,
        "show_closed": show_closed,
    }
    return render(request, "portal/usage_vendors.html", context)


@login_required
def usage_users(request):
    """
    User inventory, базирано на usage snapshot-а.
    """
    snapshot = _build_usage_snapshot()
    user_rows = snapshot["user_rows"]

    show_closed = (request.GET.get("show_closed") in ("1", "true", "True", "on", "yes"))

    # ако show_closed е False – можеш да филтрираш по user.is_active, ако има такъв флаг
    if not show_closed:
        filtered = []
        for row in user_rows:
            user = row.get("user")
            if hasattr(user, "is_active"):
                if user.is_active:
                    filtered.append(row)
            else:
                filtered.append(row)
        user_rows = filtered

    context = {
        "user_rows": user_rows,
        "show_closed": show_closed,
    }
    return render(request, "portal/usage_users.html", context)



# ----------
# LOGOUT HELPER
# ----------

def portal_logout(request):
    logout(request)
    return redirect("login")

@login_required
def user_detail(request, pk: int):
    # Keep backwards compatibility: redirect to inline details in users list
    show_closed = request.GET.get("show_closed", "0")
    rows = request.GET.get("rows", "50")
    page = request.GET.get("page", "1")
    return redirect(f"/en/portal/users/?page={page}&rows={rows}&show_closed={show_closed}&selected={pk}#user-details")